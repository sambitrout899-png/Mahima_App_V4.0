"""
ANM Community App — Resource Loading & Cost Model
══════════════════════════════════════════════════
Run:
    pip install openpyxl --break-system-packages
    python3 build_anm_cost_model.py

Output: ANM_Resource_Cost_Model.xlsx  (7 sheets)

Stack: ASP.NET Core .NET 8 · React + Ionic/Capacitor · PostgreSQL 16
10-week delivery + 30-day hypercare (2 HC weeks in model)
All amounts in INR (₹)

Sheets
  1. Assumptions      — rate card, margin %, project parameters
  2. Resource Plan    — week-by-week days per role (12 weeks incl. hypercare)
  3. Cost Model       — cost build-up, 50% margin, client price, phase breakdown
  4. Weekly Cost      — weekly & cumulative cost + selling price
  5. AMC Options      — post-warranty annual maintenance contract tiers
  6. Platform Costs   — monthly AWS / Twilio / Firebase / App Store costs (100K members)
  7. Scale Scenarios  — cost comparison: 1K vs 100K users, with/without Twilio Conversations
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ───────────────────────────────────────────────
NAVY   = "1B2F6E"
GOLD   = "F0A500"
WHITE  = "FFFFFF"
MUTED  = "64748B"
BLUE   = "0000FF"
BLACK  = "000000"
GREEN  = "008000"
BG_ALT = "F5F7FA"
BG_HDR = "EEF2F7"

# ── Helpers ───────────────────────────────────────────────
def F(c): return PatternFill("solid", fgColor=c)
def fnt(sz=10, bold=False, color=BLACK, name="Arial"): return Font(name=name, size=sz, bold=bold, color=color)
def aln(h="left", v="center", wrap=False, indent=0): return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def btm_border(): return Border(bottom=Side(style="thin", color="DDDDDD"))

def hdr_cell(ws, r, c, val, bg=NAVY, tc=WHITE, sz=10, bold=True, h="left", indent=1):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(sz, bold, tc)
    cl.fill = F(bg); cl.alignment = aln(h, indent=indent)

def inp(ws, r, c, val, fmt=None, bg=None, h="center"):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(color=BLUE)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

def calc(ws, r, c, val, fmt=None, tc=BLACK, bg=None, h="right", sz=10, bold=False):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(sz, bold, tc)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

def xlink(ws, r, c, val, fmt=None, h="center", bg=None):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(color=GREEN)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

MONEY = u"₹#,##0;(₹#,##0);-"   # ₹#,##0
PCT   = "0.0%;(0.0%);-"

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS
# Rate card rows 13-20  (referenced by Cost Model & Weekly Cost)
# Pricing row 25 = Margin %  (referenced by Cost Model & Weekly Cost)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"
ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 48

# Title
ws1.row_dimensions[1].height = 44
ws1.merge_cells("A1:D1")
ws1["A1"].value = "ANM Community App — Resource & Pricing Assumptions (.NET 8 + Ionic/Capacitor + PostgreSQL)"
ws1["A1"].font  = fnt(18, True, NAVY)
ws1["A1"].fill  = F(BG_HDR)
ws1["A1"].alignment = aln("left", indent=1)

# Metadata rows 3-9
meta = [
    ("Project",    "ANM Community App"),
    ("Client",     "ANM — Church & Community Platform"),
    ("Framework",  "Proprietary .NET 8 + Ionic/Capacitor + PostgreSQL 16"),
    ("Prepared by","Solution Development Team"),
    ("Date",       "June 2025"),
    ("Currency",   "INR (Indian Rupees — ₹)"),
    ("Version",    "v2.0 — 10-Week Model"),
]
for i, (k, v) in enumerate(meta):
    r = 3 + i
    ws1.row_dimensions[r].height = 17
    ws1.cell(r,1).value = k; ws1.cell(r,1).font = fnt(9, True, MUTED)
    ws1.cell(r,2).value = v; ws1.cell(r,2).font = fnt(10, False, "1A1A2E")

ws1.row_dimensions[10].height = 8

# ── Rate Card (rows 11-20) ────────────────────────────────
ws1.row_dimensions[11].height = 22
ws1.merge_cells("A11:D11")
hdr_cell(ws1, 11, 1, "  RATE CARD — Daily Rates (INR) — Indian Consulting Market Rates")

ws1.row_dimensions[12].height = 18
for c, h in [(1,"Role"),(2,"Daily Rate (₹)"),(3,"Role ID"),(4,"Notes / Scope")]:
    ws1.cell(12,c).value = h; ws1.cell(12,c).font = fnt(9,True,NAVY)
    ws1.cell(12,c).fill  = F(BG_HDR); ws1.cell(12,c).alignment = aln("center")
    ws1.cell(12,c).border = Border(bottom=Side(style="medium", color=NAVY))

# Rows 13-20 = 8 roles  ← CRITICAL: Cost Model links to B13:B20
rates = [
    ("Project Manager",         8500,  "PM",  "Sr PM; reporting, risk, ANM stakeholder comms"),
    ("Solution Architect",     12000,  "ARC", "Architecture reviews & formal sign-off gates"),
    ("UI/UX Designer",          6500,  "DES", "Figma, design system, UAT design support"),
    ("Mobile Developer (Sr.)",  9000,  "MOB", "React + Ionic/Capacitor — iOS, Android & Web via Capacitor bridges"),
    ("Backend Developer (Sr.)", 9000,  "BAK", "ASP.NET Core .NET 8 (C#), PostgreSQL via EF Core 8, REST APIs"),
    ("DevOps Engineer",         7500,  "DEV", "AWS ECS Fargate, RDS PostgreSQL, CI/CD, CloudFront, monitoring"),
    ("QA Engineer",             5500,  "QA",  "xUnit, Testcontainers, Playwright, Appium, OWASP testing"),
    ("Security Consultant",    11000,  "SEC", "OWASP pentest, audit & sign-off report"),
]
for i, (rname, rate, rid, note) in enumerate(rates):
    r = 13 + i
    ws1.row_dimensions[r].height = 18
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = rname; ws1.cell(r,1).font = fnt(10,True,NAVY);  ws1.cell(r,1).alignment = aln()
    inp(ws1, r, 2, rate, MONEY, bg)
    ws1.cell(r,3).value = rid;   ws1.cell(r,3).font = fnt(9,False,MUTED); ws1.cell(r,3).alignment = aln("center")
    ws1.cell(r,4).value = note;  ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).alignment = aln()
    for c in range(1,5):
        ws1.cell(r,c).fill   = F(bg)
        ws1.cell(r,c).border = btm_border()

ws1.row_dimensions[21].height = 8
ws1.row_dimensions[22].height = 8

# ── Pricing Assumptions (rows 23-31) ─────────────────────
ws1.row_dimensions[23].height = 22
ws1.merge_cells("A23:D23")
hdr_cell(ws1, 23, 1, "  PRICING ASSUMPTIONS")

ws1.row_dimensions[24].height = 18
for c, h in [(1,"Parameter"),(2,"Value"),(4,"Notes")]:
    ws1.cell(24,c).value = h; ws1.cell(24,c).font = fnt(9,True,NAVY)
    ws1.cell(24,c).fill  = F(BG_HDR); ws1.cell(24,c).alignment = aln("center")
    ws1.cell(24,c).border = Border(bottom=Side(style="medium", color=NAVY))

# Row 25 = Margin %  ← referenced as Assumptions!B25 throughout
pricing_p = [
    ("Gross Margin % (on cost)",        0.50, "0.0%",   "50% margin added on top of total cost → client price"),
    ("Working Days / Week",             5,    "#,##0",   "Mon–Fri; excludes public holidays"),
    ("Hypercare Duration (weeks)",      2,    "#,##0",   "Post go-live monitoring & support period"),
    ("Sprint Length (weeks)",           2,    "#,##0",   "Agile: 2-week sprint cadence (5 sprints total)"),
    ("Project Duration — Core (wks)",  10,    "#,##0",   "Wk 1 through Wk 10 (Discovery → Go-Live)"),
    ("Total Duration incl. HC (wks)",  12,    "#,##0",   "Core 10 weeks + 2 hypercare weeks"),
]
for i, (param, val, fmt, note) in enumerate(pricing_p):
    r = 25 + i
    ws1.row_dimensions[r].height = 18
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = param; ws1.cell(r,1).font = fnt(10,True,NAVY);  ws1.cell(r,1).alignment = aln()
    inp(ws1, r, 2, val, fmt, bg)
    ws1.cell(r,4).value = note;  ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).alignment = aln()
    for c in [1,2,4]:
        ws1.cell(r,c).fill   = F(bg)
        ws1.cell(r,c).border = btm_border()

ws1.row_dimensions[32].height = 8

# ── Color legend ──────────────────────────────────────────
ws1.row_dimensions[33].height = 22
ws1.merge_cells("A33:D33")
hdr_cell(ws1, 33, 1, "  COLOR CODING LEGEND")
legend_items = [
    ("Blue text",  BLUE,  "Hardcoded input — users may change these for scenario analysis"),
    ("Black text", BLACK, "Excel formula / calculated value — do not edit manually"),
    ("Green text", GREEN, "Link pulling from another worksheet within this workbook"),
]
for i, (lbl_, tc, note) in enumerate(legend_items):
    r = 34 + i
    ws1.row_dimensions[r].height = 17
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = lbl_;       ws1.cell(r,1).font = fnt(10,True,tc);    ws1.cell(r,1).fill = F(bg)
    ws1.cell(r,2).value = "(sample)"; ws1.cell(r,2).font = fnt(10,False,tc);   ws1.cell(r,2).fill = F(bg); ws1.cell(r,2).alignment = aln("center")
    ws1.cell(r,4).value = note;       ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).fill = F(bg); ws1.cell(r,4).alignment = aln()

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — RESOURCE PLAN
# Col  A     = Role name
# Cols B–K   = Wk 1 – Wk 10   (10 project weeks)
# Cols L–M   = HC-Wk1, HC-Wk2  (hypercare)
# Col  N     = Total Days (SUM B:M)
# Col  O     = FTE Equiv (N/50)   [50 = 10 wks × 5d]
# Col  P     = Phase focus note
#
# Rows 4–11  = 8 roles
# Row  12    = Total team-days per week (SUM rows 4:11)
# Row  13    = Max capacity (40 = 8 roles × 5d)
# Row  14    = Utilisation %
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resource Plan")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B4"
ws2.column_dimensions["A"].width = 26
for ci in range(2, 14): ws2.column_dimensions[get_column_letter(ci)].width = 6
ws2.column_dimensions["N"].width = 12
ws2.column_dimensions["O"].width = 11
ws2.column_dimensions["P"].width = 34

# Title
ws2.row_dimensions[1].height = 40
ws2.merge_cells("A1:P1")
ws2["A1"].value = "ANM Community App — Resource Loading Plan  (Team-Days per Week) | .NET 8 Framework — 10-Week Delivery | INR"
ws2["A1"].font  = fnt(15, True, WHITE)
ws2["A1"].fill  = F(NAVY)
ws2["A1"].alignment = aln("left", indent=1)

# Row 2 — Phase header spans
# Phase 1: cols B-C (Wk1-2)
# Phase 2: cols D-G (Wk3-6)
# Phase 3: cols H-I (Wk7-8)
# Phase 4: cols J-K (Wk9-10)
# Hypercare: cols L-M
ws2.row_dimensions[2].height = 20
hdr_cell(ws2, 2, 1, "PHASE", NAVY, WHITE, 9, True, "center", 0)
phase_spans = [
    ("Phase 1: Discovery & Config",    2,  3,  "1D4ED8"),   # Wk1-2  → cols B-C
    ("Phase 2: Core Customisation",    4,  7,  "065F46"),   # Wk3-6  → cols D-G
    ("Phase 3: Feature Completion",    8,  9,  "92400E"),   # Wk7-8  → cols H-I
    ("Phase 4: QA, UAT & Launch",     10, 11,  "6B21A8"),   # Wk9-10 → cols J-K
    ("Hypercare (Free — 30 days)",    12, 13,  "991B1B"),   # HC1-2  → cols L-M
]
for ph, cs, ce, pc in phase_spans:
    sl, el = get_column_letter(cs), get_column_letter(ce)
    ws2.merge_cells(f"{sl}2:{el}2")
    ws2[f"{sl}2"].value = ph; ws2[f"{sl}2"].font = fnt(8,True,WHITE)
    ws2[f"{sl}2"].fill  = F(pc); ws2[f"{sl}2"].alignment = aln("center")
for c, lbl_ in [(14,"TOTAL"),(15,"FTE"),(16,"Phase Focus")]:
    ws2.cell(2,c).value = lbl_; ws2.cell(2,c).font = fnt(9,True,WHITE)
    ws2.cell(2,c).fill  = F(NAVY); ws2.cell(2,c).alignment = aln("center")

# Row 3 — Week labels
ws2.row_dimensions[3].height = 18
hdr_cell(ws2, 3, 1, "Role", NAVY, WHITE, 10, True, "left", 1)
weeks12 = ["Wk 1","Wk 2","Wk 3","Wk 4","Wk 5","Wk 6","Wk 7","Wk 8","Wk 9","Wk 10","HC-1","HC-2"]
for ci, wl in enumerate(weeks12, 2):
    ws2.cell(3,ci).value = wl; ws2.cell(3,ci).font = fnt(8,True,WHITE)
    ws2.cell(3,ci).fill  = F(NAVY); ws2.cell(3,ci).alignment = aln("center")
for c, lbl_ in [(14,"Days"),(15,"÷50d"),(16,"Notes")]:
    ws2.cell(3,c).value = lbl_; ws2.cell(3,c).font = fnt(9,True,WHITE)
    ws2.cell(3,c).fill  = F(NAVY); ws2.cell(3,c).alignment = aln("center")

# Rows 4-11 — Resource allocations
# 12 values per role: Wk1-Wk10, HC-1, HC-2
alloc = [
    # name,                          W1  W2  W3  W4  W5  W6  W7  W8  W9 W10 HC1 HC2  note
    ("Project Manager",               5,  5,  3,  3,  3,  3,  3,  3,  5,  5,  2,  2, "Discovery → Go-Live → Hypercare oversight"),
    ("Solution Architect",            5,  5,  1,  1,  0,  1,  1,  0,  1,  1,  0,  0, "Discovery & architecture; sign-off gates only"),
    ("UI/UX Designer",                5,  5,  2,  2,  1,  1,  1,  0,  0,  0,  0,  0, "Design-heavy Wks 1–4; wraps at Phase 2"),
    ("Mobile Developer (Sr.)",        0,  1,  5,  5,  5,  5,  5,  4,  3,  3,  2,  2, "Ionic/Capacitor customisation Wks 3–8; HC support"),
    ("Backend Developer (Sr.)",       2,  2,  5,  5,  5,  5,  4,  3,  3,  3,  2,  2, ".NET 8 API Wks 1–2 design; full Wks 3–8; HC support"),
    ("DevOps Engineer",               3,  3,  2,  2,  2,  2,  2,  2,  5,  5,  3,  3, "RDS/ECS setup Wks 1–2; prod deploy Wk9-10; HC monitoring"),
    ("QA Engineer",                   0,  0,  1,  1,  1,  1,  3,  4,  5,  5,  1,  1, "Engagement ramps from Wk 7; peak in QA phase Wk9-10"),
    ("Security Consultant",           0,  0,  0,  0,  0,  0,  0,  0,  4,  3,  0,  0, "OWASP pentest Wk9-10 only; written report delivered"),
]
for ri, row_data in enumerate(alloc):
    r = 4 + ri
    rname = row_data[0]; days = row_data[1:13]; note = row_data[13]
    ws2.row_dimensions[r].height = 20
    bg = WHITE if ri%2 == 0 else BG_ALT
    ws2.cell(r,1).value = rname; ws2.cell(r,1).font = fnt(10,True,NAVY)
    ws2.cell(r,1).fill  = F(bg); ws2.cell(r,1).alignment = aln()
    ws2.cell(r,1).border = btm_border()
    # Day allocations B-M (12 weeks: Wk1-10 + HC1-2)
    for ci, d in enumerate(days, 2):
        cl = ws2.cell(r,ci)
        cl.value = d if d > 0 else None
        cl.font  = fnt(10, False, BLUE)
        cl.fill  = F(bg)
        cl.alignment = aln("center")
        cl.number_format = "#,##0;[Red]-#,##0;-"
        cl.border = Border(bottom=Side(style="thin",color="DDDDDD"),
                           left =Side(style="thin",color="E0E0E0"),
                           right=Side(style="thin",color="E0E0E0"))
    # N: Total days = SUM(B:M)
    calc(ws2, r, 14, f"=SUM(B{r}:M{r})", "#,##0", BLACK, BG_HDR, "center", 10, True)
    ws2.cell(r,14).border = btm_border()
    # O: FTE equiv = total / 50  (10 weeks × 5d = 50 working days = 1 FTE)
    calc(ws2, r, 15, f"=N{r}/50", "0.0x", BLACK, BG_HDR, "center")
    ws2.cell(r,15).border = btm_border()
    # P: Notes
    ws2.cell(r,16).value = note; ws2.cell(r,16).font = fnt(9,False,MUTED)
    ws2.cell(r,16).fill  = F(bg); ws2.cell(r,16).alignment = aln()
    ws2.cell(r,16).border = btm_border()

# Row 12 — Total team-days per week
ws2.row_dimensions[12].height = 22
ws2.cell(12,1).value = "TOTAL TEAM-DAYS / WEEK"
ws2.cell(12,1).font  = fnt(10,True,WHITE); ws2.cell(12,1).fill = F(NAVY); ws2.cell(12,1).alignment = aln(indent=1)
for ci in range(2, 15):
    c = get_column_letter(ci)
    calc(ws2, 12, ci, f"=SUM({c}4:{c}11)", "#,##0", WHITE, NAVY, "center", 10, True)

# Row 13 — Max capacity (8 × 5 = 40 per week)
ws2.row_dimensions[13].height = 17
ws2.cell(13,1).value = "Max capacity (8 roles × 5d)"; ws2.cell(13,1).font = fnt(9,False,MUTED); ws2.cell(13,1).alignment = aln()
for ci in range(2,14):
    inp(ws2, 13, ci, 40, "#,##0", h="center"); ws2.cell(13,ci).font = fnt(9,False,MUTED)
calc(ws2, 13, 14, "=SUM(B13:M13)", "#,##0", MUTED, h="center")

# Row 14 — Utilisation %
ws2.row_dimensions[14].height = 17
ws2.cell(14,1).value = "Utilisation %"; ws2.cell(14,1).font = fnt(9,False,MUTED); ws2.cell(14,1).alignment = aln()
for ci in range(2,14):
    c = get_column_letter(ci)
    calc(ws2, 14, ci, f'=IFERROR({c}12/{c}13,"-")', "0%", MUTED, h="center", sz=9)

# Framework savings note
ws2.row_dimensions[15].height = 8
ws2.row_dimensions[16].height = 20
ws2.merge_cells("A16:P16")
ws2["A16"].value = "⚡  Framework Advantage: 10-week model vs 14-16 weeks greenfield — pre-built .NET 8 + Ionic/Capacitor modules (Auth, CRUD APIs, media pipeline, push notifs, SignalR) save ~4-6 weeks."
ws2["A16"].font  = fnt(9, True, "065F46")
ws2["A16"].fill  = F("D1FAE5")
ws2["A16"].alignment = aln("left", indent=1)

# Phase summary table (rows 18-27)
ws2.row_dimensions[17].height = 8
ws2.row_dimensions[18].height = 20
ws2.merge_cells("A18:G18")
hdr_cell(ws2, 18, 1, "  PHASE SUMMARY")
ws2.row_dimensions[19].height = 17
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"% of Total"),(5,"Avg daily headcount"),(6,"Framework note")]:
    ws2.cell(19,c).value = h; ws2.cell(19,c).font = fnt(9,True,NAVY)
    ws2.cell(19,c).fill  = F(BG_HDR); ws2.cell(19,c).alignment = aln("center")

ph_summ = [
    ("Phase 1: Discovery & Config",   "Wks 1–2",    "=SUM(B12:C12)",  2,  "Stakeholder workshops, DB schema, Ionic scaffold, Figma"),
    ("Phase 2: Core Customisation",   "Wks 3–6",    "=SUM(D12:G12)",  4,  "Auth, Member Dir, Prayer, Video — framework modules configured"),
    ("Phase 3: Feature Completion",   "Wks 7–8",    "=SUM(H12:I12)",  2,  "Twilio Chat/VoIP, FCM push, admin RBAC, deep-link polish"),
    ("Phase 4: QA, UAT & Launch",     "Wks 9–10",   "=SUM(J12:K12)",  2,  "xUnit + Playwright + Appium + OWASP + UAT + Store submit"),
    ("Hypercare (Free — 30 days)",  "HC 1–2",   "=SUM(L12:M12)",  2,  "Free bug fixes, monitoring, admin training, handover docs"),
    ("TOTAL",                          "12 wks",          "=SUM(B12:M12)",  0,  "10 core + 2 hypercare | Framework saves ~4–6 weeks vs greenfield"),
]
for i, (ph, wk, days_f, n_wks, framework_note) in enumerate(ph_summ):
    r = 20 + i; is_tot = (i==5)
    bg2 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT); tc2 = WHITE if is_tot else NAVY
    ws2.row_dimensions[r].height = 17
    ws2.cell(r,1).value = ph; ws2.cell(r,1).font = fnt(9,is_tot,tc2); ws2.cell(r,1).fill = F(bg2); ws2.cell(r,1).alignment = aln()
    ws2.cell(r,2).value = wk; ws2.cell(r,2).font = fnt(9,is_tot,tc2); ws2.cell(r,2).fill = F(bg2); ws2.cell(r,2).alignment = aln("center")
    calc(ws2, r, 3, days_f, "#,##0", tc2, bg2, "center", 9, is_tot)
    calc(ws2, r, 4, f'=IFERROR(C{r}/C{20+5},"-")', PCT, tc2, bg2, "center", 9)
    wk_count = n_wks if n_wks > 0 else 12
    calc(ws2, r, 5, f'=IFERROR(ROUND(C{r}/({wk_count}*5),1),"-")', "0.0", tc2, bg2, "center", 9)
    ws2.cell(r,6).value = framework_note; ws2.cell(r,6).font = fnt(8,False,tc2)
    ws2.cell(r,6).fill  = F(bg2); ws2.cell(r,6).alignment = aln()
    for c in range(1,7): ws2.cell(r,c).border = btm_border()

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — COST MODEL
# Rows 3     = headers
# Rows 4-11  = 8 roles
# Row  12    = Subtotal (total cost, no margin)
# Row  14    = Total Cost INR (= H12)
# Row  15    = Margin % (= Assumptions!B25)
# Row  16    = Margin ₹ (= H14 × H15)
# Row  17    = Client Price (= H14 + H16)
# Rows 20+   = Phase cost breakdown
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Model")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A4"
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 20
ws3.column_dimensions["G"].width = 18
ws3.column_dimensions["H"].width = 20
ws3.column_dimensions["I"].width = 11
ws3.column_dimensions["J"].width = 30

# Title
ws3.row_dimensions[1].height = 40
ws3.merge_cells("A1:J1")
ws3["A1"].value = "ANM Community App — Cost Build-up & Pricing Model | .NET 8 + Ionic/Capacitor + PostgreSQL | 10-Week Delivery | INR"
ws3["A1"].font  = fnt(15,True,WHITE); ws3["A1"].fill = F(NAVY); ws3["A1"].alignment = aln("left",indent=1)

ws3.row_dimensions[2].height = 14
ws3.merge_cells("A2:J2")
ws3["A2"].value = "Green = cross-sheet link  ·  Blue = editable input  ·  Black = formula  ·  Margin applied in rows 14–17  ·  All amounts in INR (₹)"
ws3["A2"].font  = fnt(8,False,MUTED); ws3["A2"].fill = F(BG_HDR); ws3["A2"].alignment = aln("left",indent=1)

# Column headers row 3
ws3.row_dimensions[3].height = 30
col3_hdrs = ["Role","Daily Rate (₹)","Project Days\n(Wks 1–10)","HC Days\n(2 wks)",
             "Total\nDays","Project Cost\n(₹)","HC Cost\n(₹)","Total Cost\n(₹)","% of\nTotal","Source / Notes"]
for ci, h in enumerate(col3_hdrs, 1):
    ws3.cell(3,ci).value = h; ws3.cell(3,ci).font = fnt(9,True,WHITE); ws3.cell(3,ci).fill = F(NAVY)
    ws3.cell(3,ci).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Rows 4-11 — Role cost rows
# Resource Plan:  rows 4-11, Wk1-10 = cols B:K, HC = cols L:M
# Assumptions rate rows: B13=PM...B20=Sec
role_notes = [
    "Rate: Assumptions!B13  |  Days: Resource Plan row 4, cols B:K (core) and L:M (HC)",
    "Rate: Assumptions!B14  |  Primarily Wks 1–2; sign-off gates in Wk6 & Wk10",
    "Rate: Assumptions!B15  |  Design-heavy Wks 1–4; wraps by Wk7",
    "Rate: Assumptions!B16  |  Ionic/Capacitor customisation Wks 3–8; HC support",
    "Rate: Assumptions!B17  |  ASP.NET Core .NET 8 APIs Wks 1–10; HC support",
    "Rate: Assumptions!B18  |  RDS/ECS setup Wks 1–2; prod deploy Wk9–10; HC monitoring",
    "Rate: Assumptions!B19  |  xUnit + Playwright + Appium; ramps from Wk7; peak Wk9–10",
    "Rate: Assumptions!B20  |  OWASP pentest Wk9–10 only; fixed scope; report delivered",
]
role_data_rows = []
for ri in range(8):
    r = 4 + ri
    rp_row  = 4 + ri
    ar_row  = 13 + ri
    role_data_rows.append(r)
    ws3.row_dimensions[r].height = 20
    bg = WHITE if ri%2==0 else BG_ALT
    ws3.cell(r,1).value = alloc[ri][0]; ws3.cell(r,1).font = fnt(10,True,NAVY)
    ws3.cell(r,1).fill  = F(bg); ws3.cell(r,1).alignment = aln()
    # B: Daily rate (green link to Assumptions)
    xlink(ws3, r, 2, f"=Assumptions!B{ar_row}", MONEY, "center", bg)
    # C: Project days Wks 1-10 (Resource Plan cols B:K)
    xlink(ws3, r, 3, f"=SUM('Resource Plan'!B{rp_row}:K{rp_row})", "#,##0", "center", bg)
    # D: Hypercare days (Resource Plan cols L:M)
    xlink(ws3, r, 4, f"=SUM('Resource Plan'!L{rp_row}:M{rp_row})", "#,##0", "center", bg)
    # E: Total days = C + D
    calc(ws3, r, 5, f"=C{r}+D{r}", "#,##0", BLACK, bg, "center")
    # F: Project cost = B × C
    calc(ws3, r, 6, f"=B{r}*C{r}", MONEY, BLACK, bg)
    # G: HC cost = B × D
    calc(ws3, r, 7, f"=B{r}*D{r}", MONEY, BLACK, bg)
    # H: Total cost = F + G
    calc(ws3, r, 8, f"=F{r}+G{r}", MONEY, BLACK, bg)
    # I: % of total
    ws3.cell(r,9).font = fnt(10,False,BLACK); ws3.cell(r,9).fill = F(bg)
    ws3.cell(r,9).alignment = aln("center"); ws3.cell(r,9).number_format = PCT
    # J: Note
    ws3.cell(r,10).value = role_notes[ri]; ws3.cell(r,10).font = fnt(8,False,MUTED); ws3.cell(r,10).alignment = aln()
    for c in range(1,10): ws3.cell(r,c).border = btm_border()

# Row 12 — Subtotal
SUB = 12
ws3.row_dimensions[SUB].height = 22
hdr_cell(ws3, SUB, 1, "  TOTAL COST  (before margin)", NAVY, WHITE, 10, True, "left", 1)
ws3.cell(SUB,2).fill = F(NAVY)
for ci, col in enumerate(["C","D","E","F","G","H"], 3):
    fmt_ = MONEY if ci >= 6 else "#,##0"
    h_   = "right" if ci >= 6 else "center"
    calc(ws3, SUB, ci, f"=SUM({col}4:{col}11)", fmt_, WHITE, NAVY, h_, 10, True)
ws3.cell(SUB,9).value = "100.0%"; ws3.cell(SUB,9).font = fnt(10,True,WHITE)
ws3.cell(SUB,9).fill  = F(NAVY); ws3.cell(SUB,9).alignment = aln("center")

# Fill % of total for each role row
for r in role_data_rows:
    ws3.cell(r,9).value = f"=IFERROR(H{r}/H{SUB},0)"

# ── Pricing Block (rows 14-17) ────────────────────────────
ws3.row_dimensions[13].height = 8
PSTART = 14
pricing_rows = [
    (14, "Total Cost (₹)",              f"=H{SUB}",                   F(BG_HDR),    fnt(11,True,NAVY),          MONEY,  "right"),
    (15, "Gross Margin % (50%)",          "=Assumptions!B25",           F("FEF9C3"),  fnt(11,True,BLUE),          "0.0%", "right"),
    (16, "Margin Amount (₹)",             f"=H{PSTART}*H{PSTART+1}",   F("F0FDF4"),  fnt(11,True,"065F46"),      MONEY,  "right"),
    (17, "TOTAL CLIENT PRICE (₹)",        f"=H{PSTART}+H{PSTART+2}",  F(GOLD),      fnt(13,True,NAVY),          MONEY,  "right"),
]
for (rnum, label, val, bg_fill_, txt_fnt, fmt_, h_) in pricing_rows:
    ws3.row_dimensions[rnum].height = 26
    ws3.merge_cells(f"A{rnum}:G{rnum}")
    ws3.cell(rnum,1).value = label; ws3.cell(rnum,1).font = txt_fnt
    ws3.cell(rnum,1).fill  = bg_fill_; ws3.cell(rnum,1).alignment = aln("left", indent=2)
    ws3.cell(rnum,8).value = val;   ws3.cell(rnum,8).font = txt_fnt
    ws3.cell(rnum,8).fill  = bg_fill_; ws3.cell(rnum,8).alignment = aln(h_)
    ws3.cell(rnum,8).number_format = fmt_
    ws3.cell(rnum,8).border = Border(bottom=Side(style="medium",color="AAAAAA"), top=Side(style="thin",color="CCCCCC"))

# ── INR Package note ─────────────────────────────────────
ws3.row_dimensions[18].height = 20
ws3.merge_cells("A18:J18")
ws3["A18"].value = "\U0001f4a1  Packages: Starter ₹32L  ·  Standard ₹38L (Recommended)  ·  Enterprise ₹48L  |  AMC: ₹15K-₹40K/mo  |  AWS Infra: ~₹44,825/mo (see Platform Costs sheet)"
ws3["A18"].font  = fnt(9, True, NAVY)
ws3["A18"].fill  = F("FEF3C7")
ws3["A18"].alignment = aln("left", indent=1)

# ── Phase cost breakdown (rows 20+) — 5 phases, 10 weeks ──
ws3.row_dimensions[19].height = 8
PHSTART3 = 20
ws3.row_dimensions[PHSTART3].height = 22
ws3.merge_cells(f"A{PHSTART3}:J{PHSTART3}")
hdr_cell(ws3, PHSTART3, 1, "  COST & PRICE BY DELIVERY PHASE  (10-Week + Hypercare Model)")

ws3.row_dimensions[PHSTART3+1].height = 28
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"Proportional Cost (₹)"),(5,"Selling Price incl. 50% Margin"),(6,"% of Total Price")]:
    ws3.cell(PHSTART3+1,c).value = h; ws3.cell(PHSTART3+1,c).font = fnt(9,True,NAVY)
    ws3.cell(PHSTART3+1,c).fill  = F(BG_HDR); ws3.cell(PHSTART3+1,c).alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

# Phase team-days from Resource Plan row 12
# Wk1-2 = cols B-C, Wk3-6 = cols D-G, Wk7-8 = cols H-I, Wk9-10 = cols J-K, HC1-2 = cols L-M
ph_cost_data = [
    ("Phase 1: Discovery & Config",   "Wks 1–2",   "=SUM('Resource Plan'!B12:C12)"),
    ("Phase 2: Core Customisation",   "Wks 3–6",   "=SUM('Resource Plan'!D12:G12)"),
    ("Phase 3: Feature Completion",   "Wks 7–8",   "=SUM('Resource Plan'!H12:I12)"),
    ("Phase 4: QA, UAT & Launch",     "Wks 9–10",  "=SUM('Resource Plan'!J12:K12)"),
    ("Hypercare (Free — 30 days)",  "HC 1–2",  "=SUM('Resource Plan'!L12:M12)"),
    ("TOTAL",                          "12 wks",         f"=E{SUB}"),
]
for i, (ph, wk, days_f) in enumerate(ph_cost_data):
    r = PHSTART3 + 2 + i
    ws3.row_dimensions[r].height = 18
    is_tot = (i == 5)
    bg3 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT)
    tc3 = WHITE if is_tot else NAVY
    ws3.cell(r,1).value = ph; ws3.cell(r,1).font = fnt(9,is_tot,tc3); ws3.cell(r,1).fill = F(bg3); ws3.cell(r,1).alignment = aln()
    ws3.cell(r,2).value = wk; ws3.cell(r,2).font = fnt(9,is_tot,tc3); ws3.cell(r,2).fill = F(bg3); ws3.cell(r,2).alignment = aln("center")
    if not is_tot:
        xlink(ws3, r, 3, days_f, "#,##0", "center", bg3)
    else:
        calc(ws3, r, 3, days_f, "#,##0", WHITE, NAVY, "center", 9, True)
    calc(ws3, r, 4, f"=IFERROR(C{r}/E{SUB}*H{SUB},0)", MONEY, tc3, bg3, "right", 9, is_tot)
    sell_tc = GOLD if is_tot else "065F46"; sell_bg = GOLD if is_tot else bg3
    calc(ws3, r, 5, f"=IFERROR(D{r}*(1+Assumptions!B25),0)", MONEY, sell_tc, sell_bg, "right", 9, is_tot)
    client_price_row = PSTART + 3
    calc(ws3, r, 6, f'=IFERROR(E{r}/H{client_price_row},"-")', PCT, tc3, bg3, "center", 9)
    for c in range(1,7): ws3.cell(r,c).border = btm_border()

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — WEEKLY COST
# Row  3     = headers
# Rows 4-15  = 12 weeks (Wk1–Wk10, HC-Wk1, HC-Wk2)
# Row  16    = TOTAL
# Rows 18+   = Phase summary
#
# Cols: A=Week, B=Phase, C-J=8 role costs, K=Total/wk,
#       L=Cumul Cost, M=Sell Price/wk, N=Cumul Sell Price
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Weekly Cost")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "C4"
ws4.column_dimensions["A"].width = 9
ws4.column_dimensions["B"].width = 24
for ci in range(3, 11): ws4.column_dimensions[get_column_letter(ci)].width = 14
ws4.column_dimensions["K"].width = 17
ws4.column_dimensions["L"].width = 18
ws4.column_dimensions["M"].width = 17
ws4.column_dimensions["N"].width = 19

# Title
ws4.row_dimensions[1].height = 40
ws4.merge_cells("A1:N1")
ws4["A1"].value = "ANM Community App — Weekly Cost Breakdown & Cash Flow | .NET 8 Framework | 10-Week Delivery | INR"
ws4["A1"].font  = fnt(15,True,WHITE); ws4["A1"].fill = F(NAVY); ws4["A1"].alignment = aln("left",indent=1)

ws4.row_dimensions[2].height = 14
ws4.merge_cells("A2:N2")
ws4["A2"].value = "Costs = days (Resource Plan) × rate (Assumptions)  ·  Selling price = cost × (1 + 50% margin)  ·  All amounts in INR (₹)  ·  12 weeks total incl. 2 hypercare"
ws4["A2"].font  = fnt(8,False,MUTED); ws4["A2"].fill = F(BG_HDR); ws4["A2"].alignment = aln("left",indent=1)

# Column headers row 3
ws4.row_dimensions[3].height = 30
role_short = ["PM","Architect","Designer","Mobile Dev\n(Ionic/Cap)","Backend Dev\n(.NET 8)","DevOps\n(RDS/ECS)","QA Eng","Security"]
col4_hdrs  = ["Week","Phase"] + role_short + ["Total Cost\n/Wk (₹)","Cumul. Cost\n(₹)","Sell Price\n/Wk (₹)","Cumul. Sell\nPrice (₹)"]
for ci, h in enumerate(col4_hdrs, 1):
    ws4.cell(3,ci).value = h; ws4.cell(3,ci).font = fnt(9,True,WHITE); ws4.cell(3,ci).fill = F(NAVY)
    ws4.cell(3,ci).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wk_phase_info = [
    ("Wk 1",   "Phase 1: Discovery & Config",  "EFF6FF"),
    ("Wk 2",   "Phase 1: Discovery & Config",  "EFF6FF"),
    ("Wk 3",   "Phase 2: Core Customisation",  "F0FDF4"),
    ("Wk 4",   "Phase 2: Core Customisation",  "F0FDF4"),
    ("Wk 5",   "Phase 2: Core Customisation",  "F0FDF4"),
    ("Wk 6",   "Phase 2: Core Customisation",  "F0FDF4"),
    ("Wk 7",   "Phase 3: Feature Completion",  "FFFBEB"),
    ("Wk 8",   "Phase 3: Feature Completion",  "FFFBEB"),
    ("Wk 9",   "Phase 4: QA, UAT & Launch",    "F5F3FF"),
    ("Wk 10",  "Phase 4: QA, UAT & Launch",    "F5F3FF"),
    ("HC-Wk1", "Hypercare (Free — 30 days)", "FFF1F2"),
    ("HC-Wk2", "Hypercare (Free — 30 days)", "FFF1F2"),
]
# Resource Plan: Wk1=B, Wk2=C, Wk3=D, Wk4=E, Wk5=F, Wk6=G, Wk7=H, Wk8=I, Wk9=J, Wk10=K, HC1=L, HC2=M
rp_week_cols = ["B","C","D","E","F","G","H","I","J","K","L","M"]
rp_rows4 = [4,5,6,7,8,9,10,11]
assum_rates = [f"Assumptions!B{13+i}" for i in range(8)]

for wi, (wlbl, wphase, wbg) in enumerate(wk_phase_info):
    r = 4 + wi
    ws4.row_dimensions[r].height = 20
    rp_col = rp_week_cols[wi]
    ws4.cell(r,1).value = wlbl;   ws4.cell(r,1).font = fnt(10,True,NAVY);  ws4.cell(r,1).fill = F(wbg); ws4.cell(r,1).alignment = aln("center")
    ws4.cell(r,2).value = wphase; ws4.cell(r,2).font = fnt(9,False,MUTED); ws4.cell(r,2).fill = F(wbg); ws4.cell(r,2).alignment = aln()
    for ci_off, (rp_row, rate_ref) in enumerate(zip(rp_rows4, assum_rates)):
        ci = 3 + ci_off
        calc(ws4, r, ci, f"='Resource Plan'!{rp_col}{rp_row}*{rate_ref}", MONEY, BLACK, wbg)
        ws4.cell(r,ci).border = btm_border()
    calc(ws4, r, 11, f"=SUM(C{r}:J{r})", MONEY, BLACK, BG_HDR, "right", 10, True)
    ws4.cell(r,11).border = Border(bottom=Side(style="thin",color="DDDDDD"), left=Side(style="medium",color=NAVY))
    if wi == 0:
        calc(ws4, r, 12, f"=K{r}", MONEY, BLACK, BG_HDR, "right")
    else:
        calc(ws4, r, 12, f"=L{r-1}+K{r}", MONEY, BLACK, BG_HDR, "right")
    ws4.cell(r,12).border = btm_border()
    calc(ws4, r, 13, f"=K{r}*(1+Assumptions!B25)", MONEY, "065F46", "F0FDF4", "right", 10, True)
    ws4.cell(r,13).border = Border(bottom=Side(style="thin",color="DDDDDD"), left=Side(style="medium",color=NAVY))
    if wi == 0:
        calc(ws4, r, 14, f"=M{r}", MONEY, "065F46", "F0FDF4", "right")
    else:
        calc(ws4, r, 14, f"=N{r-1}+M{r}", MONEY, "065F46", "F0FDF4", "right")
    ws4.cell(r,14).border = btm_border()

# Row 16 — Grand total (12 weeks: rows 4-15)
GTROW = 16
ws4.row_dimensions[GTROW].height = 24
ws4.cell(GTROW,1).value = "TOTAL"; ws4.cell(GTROW,1).font = fnt(11,True,WHITE); ws4.cell(GTROW,1).fill = F(NAVY); ws4.cell(GTROW,1).alignment = aln("center")
ws4.cell(GTROW,2).fill = F(NAVY)
for ci in range(3, 15):
    c = get_column_letter(ci)
    if ci in [12, 14]:
        calc(ws4, GTROW, ci, f"={c}{GTROW-1}", MONEY, NAVY, GOLD, "right", 12, True)
    else:
        calc(ws4, GTROW, ci, f"=SUM({c}4:{c}{GTROW-1})", MONEY, WHITE, NAVY, "right", 11, True)

# Phase summary (rows 18+)
ws4.row_dimensions[17].height = 8
PHROW4 = 18
ws4.row_dimensions[PHROW4].height = 22
ws4.merge_cells(f"A{PHROW4}:N{PHROW4}")
hdr_cell(ws4, PHROW4, 1, "  COST SUMMARY BY PHASE  (10-Week + Hypercare Model)")
ws4.row_dimensions[PHROW4+1].height = 18
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"Cost (₹)"),(5,"Selling Price (₹)"),(6,"% of Total")]:
    ws4.cell(PHROW4+1,c).value = h; ws4.cell(PHROW4+1,c).font = fnt(9,True,NAVY)
    ws4.cell(PHROW4+1,c).fill  = F(BG_HDR); ws4.cell(PHROW4+1,c).alignment = aln("center")

# Wk phase ranges in Weekly Cost rows 4-15
ph4_ranges = [
    ("Phase 1: Discovery & Config",    "Wks 1–2",   "K4:K5",    "M4:M5",    "=SUM('Resource Plan'!B12:C12)"),
    ("Phase 2: Core Customisation",    "Wks 3–6",   "K6:K9",    "M6:M9",    "=SUM('Resource Plan'!D12:G12)"),
    ("Phase 3: Feature Completion",    "Wks 7–8",   "K10:K11",  "M10:M11",  "=SUM('Resource Plan'!H12:I12)"),
    ("Phase 4: QA, UAT & Launch",      "Wks 9–10",  "K12:K13",  "M12:M13",  "=SUM('Resource Plan'!J12:K12)"),
    ("Hypercare (Free — 30 days)", "HC 1–2",   "K14:K15",  "M14:M15",  "=SUM('Resource Plan'!L12:M12)"),
    ("TOTAL",                           "12 wks",         "K4:K15",   "M4:M15",   "=SUM('Resource Plan'!B12:M12)"),
]
TOTAL_SELL_ROW = PHROW4 + 2 + 5
for i, (ph, wk, kr, mr, days_f) in enumerate(ph4_ranges):
    r = PHROW4 + 2 + i
    ws4.row_dimensions[r].height = 18
    is_tot = (i == 5)
    bg4 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT); tc4 = WHITE if is_tot else NAVY
    ws4.cell(r,1).value = ph; ws4.cell(r,1).font = fnt(9,is_tot,tc4); ws4.cell(r,1).fill = F(bg4); ws4.cell(r,1).alignment = aln()
    ws4.cell(r,2).value = wk; ws4.cell(r,2).font = fnt(9,is_tot,tc4); ws4.cell(r,2).fill = F(bg4); ws4.cell(r,2).alignment = aln("center")
    xlink(ws4, r, 3, days_f, "#,##0", "center", bg4) if not is_tot else calc(ws4, r, 3, days_f, "#,##0", WHITE, NAVY, "center", 9, True)
    calc(ws4, r, 4, f"=SUM({kr})", MONEY, tc4, bg4, "right", 9, is_tot)
    calc(ws4, r, 5, f"=SUM({mr})", MONEY, GOLD if is_tot else "065F46", GOLD if is_tot else bg4, "right", 9, is_tot)
    calc(ws4, r, 6, f'=IFERROR(E{r}/E{TOTAL_SELL_ROW},"-")', PCT, tc4, bg4, "center", 9)
    for c in range(1,7): ws4.cell(r,c).border = btm_border()

# ═══════════════════════════════════════════════════════════════
# SHEET 5 — AMC OPTIONS
# Annual Maintenance Contract tiers post go-live
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("AMC Options")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 20
ws5.column_dimensions["D"].width = 20
ws5.column_dimensions["E"].width = 26

# Title
ws5.row_dimensions[1].height = 44
ws5.merge_cells("A1:E1")
ws5["A1"].value = "ANM Community App — Annual Maintenance Contract (AMC) Options | Post Go-Live | INR"
ws5["A1"].font  = fnt(17,True,WHITE); ws5["A1"].fill = F(NAVY); ws5["A1"].alignment = aln("left",indent=1)

ws5.row_dimensions[2].height = 14
ws5.merge_cells("A2:E2")
ws5["A2"].value = "Choose an AMC tier based on desired SLA, feature development needs, and budget. AMC pricing is per month and billed quarterly."
ws5["A2"].font  = fnt(9,False,MUTED); ws5["A2"].fill = F(BG_HDR); ws5["A2"].alignment = aln("left",indent=1)

# Header row 4
ws5.row_dimensions[4].height = 22
for c, h in [(1,"Feature / Parameter"),(2,"Basic"),(3,"⭐ Standard (Recommended)"),(4,"Premium"),(5,"Notes")]:
    ws5.cell(4,c).value = h; ws5.cell(4,c).font = fnt(10,True,WHITE)
    ws5.cell(4,c).fill  = F(NAVY if c!=3 else "1D4ED8"); ws5.cell(4,c).alignment = aln("center")
    ws5.cell(4,c).border = Border(bottom=Side(style="medium",color=GOLD))

# Monthly pricing row (row 5)
ws5.row_dimensions[5].height = 30
ws5.cell(5,1).value = "Monthly Price (₹)"; ws5.cell(5,1).font = fnt(11,True,NAVY); ws5.cell(5,1).fill = F(BG_HDR); ws5.cell(5,1).alignment = aln()
amc_prices = [15000, 25000, 40000]
amc_colors = ["64748B", "1B2F6E", "92400E"]
for ci, (price, col) in enumerate(zip(amc_prices, amc_colors), 2):
    ws5.cell(5,ci).value = price; ws5.cell(5,ci).font = fnt(14,True,col)
    ws5.cell(5,ci).number_format = MONEY; ws5.cell(5,ci).alignment = aln("center")
    ws5.cell(5,ci).fill = F("F8FAFC" if ci==2 else ("EFF6FF" if ci==3 else "FEF3C7"))
    ws5.cell(5,ci).border = Border(bottom=Side(style="medium",color=GOLD))
ws5.cell(5,5).value = "Billed quarterly in advance"; ws5.cell(5,5).font = fnt(9,False,MUTED); ws5.cell(5,5).alignment = aln()

# Annual cost row (row 6)
ws5.row_dimensions[6].height = 20
ws5.cell(6,1).value = "Annual Cost (₹)"; ws5.cell(6,1).font = fnt(9,True,MUTED); ws5.cell(6,1).alignment = aln(indent=1)
for ci in range(2,5):
    ws5.cell(6,ci).value = f"=B5*12" if ci==2 else (f"=C5*12" if ci==3 else f"=D5*12")
    ws5.cell(6,ci).font = fnt(9,False,MUTED); ws5.cell(6,ci).alignment = aln("center")
    ws5.cell(6,ci).number_format = MONEY

# AMC features
amc_features = [
    ("Bug Fixes & Patches",                "Included",     "Included",     "Included",     "OS & device compatibility + security patches"),
    ("Monthly Health Check Report",        "Included",     "Included",     "Included",     "Performance, uptime, DB metrics report"),
    ("Response SLA",                       "48 hours",     "8 hours",      "2 hours",      "From ticket raised to first response"),
    ("Resolution SLA (critical bugs)",     "5 business days","2 business days","Next business day","P1 critical issue resolution"),
    ("Minor UI/UX Enhancements",           "–",       "2 per quarter","Unlimited*",   "*Subject to 4h/month cap for Standard"),
    ("Feature Development (hours/mo)",     "0 hours",      "8 hours",      "20 hours",     "Pre-agreed scope; billed additionally beyond cap"),
    ("AWS Cost Optimisation Review",       "–",       "Quarterly",    "Monthly",      "RDS, ECS, S3 cost right-sizing"),
    ("Dedicated Support Manager",          "–",       "–",       "Included",     "Named contact for all support escalations"),
    ("Priority Queue",                     "–",       "Yes",          "Yes (highest)","Tickets prioritised above non-AMC clients"),
    ("App Store Update Submissions",       "1 per quarter","2 per quarter","Unlimited",    "Binary + metadata updates to Apple & Google"),
    ("Database Backup Verification",       "Monthly",      "Weekly",       "Daily",        "RDS automated backup restore test"),
    ("Uptime SLA",                         "–",       "99.5%",        "99.9%",        "Based on AWS CloudWatch metrics"),
    ("Quarterly Business Review",          "–",       "Included",     "Included",     "30-min roadmap + usage + cost review call"),
    ("Phase 2 Feature Backlog Grooming",   "–",       "–",       "Included",     "Monthly 1h session to refine enhancement backlog"),
]
for i, (feat, basic, std, prem, note) in enumerate(amc_features):
    r = 8 + i
    ws5.row_dimensions[r].height = 18
    bg = WHITE if i%2==0 else BG_ALT
    ws5.cell(r,1).value = feat;  ws5.cell(r,1).font = fnt(9,True,NAVY);  ws5.cell(r,1).fill = F(bg);  ws5.cell(r,1).alignment = aln(indent=1)
    ws5.cell(r,2).value = basic; ws5.cell(r,2).font = fnt(9,False,MUTED if basic=="–" else BLACK); ws5.cell(r,2).fill = F(bg); ws5.cell(r,2).alignment = aln("center")
    ws5.cell(r,3).value = std;   ws5.cell(r,3).font = fnt(9,True if std!="–" else False,"1D4ED8" if std!="–" else MUTED); ws5.cell(r,3).fill = F("EFF6FF"); ws5.cell(r,3).alignment = aln("center")
    ws5.cell(r,4).value = prem;  ws5.cell(r,4).font = fnt(9,True if prem!="–" else False,"065F46" if prem!="–" else MUTED); ws5.cell(r,4).fill = F(bg); ws5.cell(r,4).alignment = aln("center")
    ws5.cell(r,5).value = note;  ws5.cell(r,5).font = fnt(8,False,MUTED); ws5.cell(r,5).fill = F(bg); ws5.cell(r,5).alignment = aln()
    for c in range(1,6): ws5.cell(r,c).border = btm_border()

# Note row
ws5.row_dimensions[23].height = 8
ws5.row_dimensions[24].height = 20
ws5.merge_cells("A24:E24")
ws5["A24"].value = "⚠️  AMC begins after the 30-day free Hypercare period. All AMC tiers include Twilio & Firebase config updates. AWS costs are pass-through (billed separately)."
ws5["A24"].font  = fnt(9,True,NAVY); ws5["A24"].fill = F("FEF3C7"); ws5["A24"].alignment = aln("left",indent=1)

# ═══════════════════════════════════════════════════════════════
# SHEET 6 — PLATFORM & INFRASTRUCTURE COSTS
# Monthly AWS, Twilio, Firebase, App Store costs — client-borne
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Platform Costs")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 32
ws6.column_dimensions["B"].width = 18
ws6.column_dimensions["C"].width = 18
ws6.column_dimensions["D"].width = 14
ws6.column_dimensions["E"].width = 34

# Title
ws6.row_dimensions[1].height = 44
ws6.merge_cells("A1:E1")
ws6["A1"].value = "ANM Community App — Monthly Platform & Infrastructure Costs | Client-Borne | INR"
ws6["A1"].font  = fnt(17,True,WHITE); ws6["A1"].fill = F(NAVY); ws6["A1"].alignment = aln("left",indent=1)

ws6.row_dimensions[2].height = 14
ws6.merge_cells("A2:E2")
ws6["A2"].value = "These costs are client-borne and billed directly by AWS / Twilio / Firebase. They are OVER AND ABOVE the implementation fee. Exchange rate basis: 1 USD ≈ ₹84. Sized for 100,000 church members."
ws6["A2"].font  = fnt(9,False,MUTED); ws6["A2"].fill = F(BG_HDR); ws6["A2"].alignment = aln("left",indent=1)

platform_sections = [
    # (section_title, color, items)
    # Sized for 100,000 church members (Low = ~15-20% MAU active; High = ~35-40% MAU active)
    ("AWS — Compute & Database", "1D4ED8", [
        ("AWS RDS PostgreSQL 16  (db.r6g.large, Multi-AZ, 500 GB gp3)",  30000, 30000, "Monthly",  "Upgraded from t4g.medium — 2 vCPU, 16 GB RAM required at 100K member scale"),
        ("AWS ECS Fargate  (4 baseline tasks / 2 vCPU / 4 GB, scales to 15)", 28000, 45000, "Monthly", ".NET 8 containers; auto-scale 4–15 tasks; avg 5–8 tasks in practice"),
        ("AWS S3 + CloudFront CDN  (1 TB storage + 3–7 TB egress/mo)",   20000, 40000, "Monthly",  "Video library; 100K members watching sermons = 3–7 TB CDN egress/mo"),
        ("AWS CloudWatch + X-Ray  (metrics, traces, logs)",                3500,  5000,  "Monthly",  "Monitoring, alerting, distributed tracing — higher volume at 100K scale"),
        ("AWS ALB + WAF  (managed rule groups, ~10M req/mo)",              4500,  6000,  "Monthly",  "WAF managed rules ($20/mo extra) + more LCUs at higher traffic"),
        ("AWS Route 53  (DNS hosting + health checks)",                    500,   500,   "Monthly",  "SSL via ACM (free) + DNS zone + health check routing"),
        ("AWS ElastiCache Redis  (cache.t3.small, 2-node HA cluster)",    4200,  4200,  "Monthly",  "NEW at 100K — session cache, query cache, rate-limit counters; essential for performance"),
        ("AWS NAT Gateway  (private subnet for ECS tasks)",               3000,  4500,  "Monthly",  "NEW — required for ECS in private VPC subnet; ~50–100 GB data processed/mo"),
        ("AWS SES  (transactional email, est. 50,000 msgs/mo)",           500,   800,   "Monthly",  "Welcome, digest, alerts — $0.10/1,000 emails via SES"),
    ]),
    ("Twilio — Communications", "059669", [
        ("Twilio Voice  (in-app VoIP, est. 2,000–5,000 mins/mo)",        8400,  21000, "Monthly",  "~₹4.2/min; scales with call adoption. 2K min low / 5K min high"),
        ("Twilio SMS  (OTP + notifications, est. 1,500–3,000 msgs/mo)",  7350,  14700, "Monthly",  "~₹4.9/SMS; India DLT registered; 1,500 low / 3,000 high msgs/mo"),
        ("⚠️ Twilio Conversations  (Chat SDK — 10K–20K active MAU)",      42000, 84000, "Monthly",  "CRITICAL: $0.05/active chat user/mo. 10% adoption=10K MAU=₹42K; 20%=₹84K. See Scale Scenarios sheet for alternative."),
    ]),
    ("Firebase — Auth & Notifications", "D97706", [
        ("Firebase Auth  (email/password + Google OAuth)",                 0,     0,     "Free",     "Email/password & Google OAuth: free at any scale — no per-user cost"),
        ("Firebase Auth  (phone OTP — new signups only, ~1,000/mo)",      504,   1008,  "Monthly",  "$0.006/SMS verification; only new monthly sign-ups trigger this cost"),
        ("Firebase FCM Push Notifications",                                0,     0,     "Free",     "Unlimited push notifications — no cost regardless of scale"),
        ("Firebase Hosting  (PWA served via CloudFront origin shield)",    0,     0,     "Free",     "CloudFront absorbs egress; Firebase serves minimal traffic to CDN origin"),
    ]),
    ("App Stores & Third-Party", "7C3AED", [
        ("Apple Developer Program  (annual — amortised monthly)",          692,   692,   "Monthly",  "₹8,299/yr ÷ 12 = ₹692/mo shown here; billed annually by Apple"),
        ("Google Play Console  (one-time registration)",                   2500,  2500,  "One-time", "₹2,500 one-time; no annual fee"),
        ("Twilio DLT Registration  (India TRAI — one-time)",               7500,  10000, "One-time", "Required for transactional SMS in India; process takes 1–2 weeks"),
    ]),
]

cur_row = 4
GRAND_MONTHLY_SUM = []
for section_title, sec_color, items in platform_sections:
    ws6.row_dimensions[cur_row].height = 22
    ws6.merge_cells(f"A{cur_row}:E{cur_row}")
    hdr_cell(ws6, cur_row, 1, f"  {section_title}", sec_color, WHITE, 10, True, "left", 1)
    cur_row += 1

    ws6.row_dimensions[cur_row].height = 18
    for c, h in [(1,"Service / Resource"),(2,"Cost Low (₹)"),(3,"Cost High (₹)"),(4,"Frequency"),(5,"Notes")]:
        ws6.cell(cur_row,c).value = h; ws6.cell(cur_row,c).font = fnt(9,True,NAVY)
        ws6.cell(cur_row,c).fill  = F(BG_HDR); ws6.cell(cur_row,c).alignment = aln("center")
        ws6.cell(cur_row,c).border = Border(bottom=Side(style="medium",color=NAVY))
    cur_row += 1

    for j, (svc, low, high, freq, note) in enumerate(items):
        ws6.row_dimensions[cur_row].height = 18
        bg = WHITE if j%2==0 else BG_ALT
        ws6.cell(cur_row,1).value = svc;  ws6.cell(cur_row,1).font = fnt(9,False,NAVY);  ws6.cell(cur_row,1).fill = F(bg); ws6.cell(cur_row,1).alignment = aln(indent=1)
        if low == 0:
            ws6.cell(cur_row,2).value = "Free"; ws6.cell(cur_row,2).font = fnt(9,True,"065F46"); ws6.cell(cur_row,2).alignment = aln("center")
            ws6.cell(cur_row,3).value = "Free"; ws6.cell(cur_row,3).font = fnt(9,True,"065F46"); ws6.cell(cur_row,3).alignment = aln("center")
        else:
            inp(ws6, cur_row, 2, low, MONEY, bg)
            inp(ws6, cur_row, 3, high, MONEY, bg)
            if freq == "Monthly":
                GRAND_MONTHLY_SUM.append(f"B{cur_row}")
        ws6.cell(cur_row,2).fill = F(bg); ws6.cell(cur_row,3).fill = F(bg)
        ws6.cell(cur_row,4).value = freq; ws6.cell(cur_row,4).font = fnt(9,False,MUTED); ws6.cell(cur_row,4).alignment = aln("center"); ws6.cell(cur_row,4).fill = F(bg)
        ws6.cell(cur_row,5).value = note; ws6.cell(cur_row,5).font = fnt(8,False,MUTED); ws6.cell(cur_row,5).alignment = aln(); ws6.cell(cur_row,5).fill = F(bg)
        for c in range(1,6): ws6.cell(cur_row,c).border = btm_border()
        cur_row += 1

    cur_row += 1  # gap between sections

# Grand total row
ws6.row_dimensions[cur_row].height = 26
ws6.merge_cells(f"A{cur_row}:A{cur_row}")
ws6.cell(cur_row,1).value = "ESTIMATED TOTAL MONTHLY PLATFORM COST"; ws6.cell(cur_row,1).font = fnt(11,True,WHITE); ws6.cell(cur_row,1).fill = F(NAVY); ws6.cell(cur_row,1).alignment = aln(indent=1)
total_formula = "=" + "+".join(GRAND_MONTHLY_SUM)
ws6.cell(cur_row,2).value = total_formula; ws6.cell(cur_row,2).font = fnt(12,True,GOLD); ws6.cell(cur_row,2).fill = F(NAVY); ws6.cell(cur_row,2).number_format = MONEY; ws6.cell(cur_row,2).alignment = aln("right")
ws6.cell(cur_row,3).value = total_formula.replace("B","C"); ws6.cell(cur_row,3).font = fnt(12,True,GOLD); ws6.cell(cur_row,3).fill = F(NAVY); ws6.cell(cur_row,3).number_format = MONEY; ws6.cell(cur_row,3).alignment = aln("right")
ws6.cell(cur_row,4).value = "Monthly"; ws6.cell(cur_row,4).font = fnt(10,True,WHITE); ws6.cell(cur_row,4).fill = F(NAVY); ws6.cell(cur_row,4).alignment = aln("center")
ws6.cell(cur_row,5).value = "Monthly recurring only. Excl. one-time fees (Google Play ₹2,500 + Twilio DLT ₹7,500–10,000) + Apple Dev ₹692/mo amortised. See Scale Scenarios sheet."; ws6.cell(cur_row,5).font = fnt(8,False,"FFFFFFBB"); ws6.cell(cur_row,5).fill = F(NAVY); ws6.cell(cur_row,5).alignment = aln()

cur_row += 2
ws6.row_dimensions[cur_row].height = 20
ws6.merge_cells(f"A{cur_row}:E{cur_row}")
ws6[f"A{cur_row}"].value = "\U0001f4a1  Sized for 100,000 church members (Low = ~15–20% MAU active; High = ~35–40% MAU active). ⚠️  Twilio Conversations at ₹42K–84K/mo is the #1 cost driver — see Scale Scenarios sheet for SignalR alternative (saves ₹42K–84K/mo). AWS costs vary with usage; review quarterly."
ws6[f"A{cur_row}"].font  = fnt(9,True,NAVY); ws6[f"A{cur_row}"].fill = F("FEF3C7"); ws6[f"A{cur_row}"].alignment = aln("left",indent=1)

# ═══════════════════════════════════════════════════════════════
# SHEET 7 — SCALE SCENARIOS
# Comparison: 500-1K users vs 100K users
# With Twilio Conversations vs Without (SignalR in-stack alternative)
# ═══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Scale Scenarios")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 34
ws7.column_dimensions["B"].width = 16
ws7.column_dimensions["C"].width = 16
ws7.column_dimensions["D"].width = 16
ws7.column_dimensions["E"].width = 16
ws7.column_dimensions["F"].width = 32

# Title
ws7.row_dimensions[1].height = 44
ws7.merge_cells("A1:F1")
ws7["A1"].value = "ANM Community App — Monthly Platform Cost: Scale Scenarios Comparison | INR"
ws7["A1"].font  = fnt(17,True,WHITE); ws7["A1"].fill = F(NAVY); ws7["A1"].alignment = aln("left",indent=1)

ws7.row_dimensions[2].height = 14
ws7.merge_cells("A2:F2")
ws7["A2"].value = "Exchange rate: 1 USD ≈ ₹84 | Low = conservative MAU activity | High = peak MAU activity | ⚠️ Twilio Conversations = $0.05/active chat user/month"
ws7["A2"].font  = fnt(9,False,MUTED); ws7["A2"].fill = F(BG_HDR); ws7["A2"].alignment = aln("left",indent=1)

# Column headers row 4
ws7.row_dimensions[4].height = 28
scenario_headers = [
    (1, "Service / Cost Driver",        NAVY,  WHITE, "left"),
    (2, "500–1K Users\nLow (₹)",        "1D4ED8", WHITE, "center"),
    (3, "500–1K Users\nHigh (₹)",       "1D4ED8", WHITE, "center"),
    (4, "100K Users\nLow (₹)",          "991B1B", WHITE, "center"),
    (5, "100K Users\nHigh (₹)",         "991B1B", WHITE, "center"),
    (6, "Notes",                         NAVY,  WHITE, "left"),
]
for col, hdr, bg, tc, ha in scenario_headers:
    ws7.cell(4,col).value = hdr
    ws7.cell(4,col).font = fnt(9,True,tc)
    ws7.cell(4,col).fill = F(bg)
    ws7.cell(4,col).alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    ws7.cell(4,col).border = Border(bottom=Side(style="medium",color=GOLD))

# Scenario rows — (label, s1_low, s1_high, s100_low, s100_high, note, is_section, is_warning, is_total)
# is_section = True → blue section banner
# is_warning = True → amber highlight
# is_total   = True → dark navy total row
scenario_rows = [
    # ── AWS ──
    ("AWS — COMPUTE & DATABASE", 0,0,0,0,"",True,False,False),
    ("RDS PostgreSQL",           18500, 18500, 30000, 30000, "Upgraded: t4g.medium → r6g.large + 500 GB storage at 100K scale",False,False,False),
    ("ECS Fargate",              12000, 12000, 28000, 45000, "4–15 tasks (2 vCPU/4 GB) vs 2 tasks (1 vCPU/2 GB) at small scale",False,False,False),
    ("S3 + CloudFront CDN",       6000,  6000, 20000, 40000, "3–7 TB CDN egress at 100K vs 1 TB at small scale",False,False,False),
    ("CloudWatch + X-Ray",        1500,  1500,  3500,  5000, "Higher log volume and trace data at 100K",False,False,False),
    ("ALB + WAF",                 2000,  2000,  4500,  6000, "WAF managed rules added at 100K; more LCUs",False,False,False),
    ("Route 53",                   500,   500,   500,   500, "Unchanged",False,False,False),
    ("ElastiCache Redis",            0,     0,  4200,  4200, "NEW at 100K — caching essential for performance at scale",False,True,False),
    ("NAT Gateway",                  0,     0,  3000,  4500, "NEW at 100K — required for ECS in private VPC subnet",False,True,False),
    ("AWS SES (email)",            800,   800,   500,   800, "Scaled email volume; SES replaces SendGrid at 100K",False,False,False),
    ("AWS Subtotal",             41300, 41300, 94200,136000, "8 services at 500-1K scale → 9 services at 100K",False,False,True),
    # ── Twilio ──
    ("TWILIO — COMMUNICATIONS", 0,0,0,0,"",True,False,False),
    ("Twilio Voice (VoIP)",       2100,  2100,  8400, 21000, "2K–5K min/mo at 100K vs 500 min/mo at small scale",False,False,False),
    ("Twilio SMS (OTP)",          1225,  1225,  7350, 14700, "1.5K–3K msgs/mo at 100K vs 250 at small scale",False,False,False),
    ("⚠️ Twilio Conversations",   1500,  1500, 42000, 84000, "⚠️ CRITICAL: $0.05/MAU. 10K MAU=₹42K; 20K MAU=₹84K. Biggest cost driver at 100K",False,True,False),
    ("Twilio Subtotal",           4825,  4825, 57750,119700, "Twilio Conversations dominates at scale",False,False,True),
    # ── Firebase ──
    ("FIREBASE — AUTH & NOTIFICATIONS", 0,0,0,0,"",True,False,False),
    ("Firebase Auth (email/Google)", 0,  0,    0,    0, "Free at any scale for email/password + Google OAuth",False,False,False),
    ("Firebase Auth (phone OTP)",    0,  0,  504, 1008, "New at 100K — ~1,000 phone OTP verifications/mo for new signups",False,False,False),
    ("Firebase FCM (push)",          0,  0,    0,    0, "Free at any scale — unlimited push notifications",False,False,False),
    ("Firebase Subtotal",            0,  0,  504, 1008, "Essentially free even at 100K if using email/Google auth",False,False,True),
    # ── Totals WITH Twilio Conversations ──
    ("SCENARIO A — WITH Twilio Conversations", 0,0,0,0,"",True,False,False),
    ("Total Monthly (with Twilio Chat)", 46125, 46125, 152454, 256708, "Full Twilio stack — Conversations SDK for in-app chat",False,False,True),
    # ── Totals WITHOUT Twilio Conversations ──
    ("SCENARIO B — WITHOUT Twilio Conversations (SignalR + PostgreSQL)", 0,0,0,0,"",True,False,False),
    ("SignalR + PostgreSQL Chat",       0,  0,  0,  0, "Built into .NET 8 + existing RDS — zero additional monthly cost",False,False,False),
    ("Total Monthly (SignalR chat)", 44625, 44625, 110454, 172708, "Recommended at 100K — saves ₹42K–84K/mo vs Twilio Conversations",False,False,True),
    ("Monthly Saving vs Scenario A", 1500, 1500, 42000, 84000, "Switch from Twilio Conversations to SignalR saves this amount per month",False,True,False),
    ("Annual Saving vs Scenario A", 18000, 18000, 504000, 1008000, "₹5.04 lakh – ₹10.08 lakh per year saved by using SignalR for chat",False,True,False),
]

r7 = 6
for row_data in scenario_rows:
    label, s1l, s1h, s100l, s100h, note, is_section, is_warn, is_total = row_data
    ws7.row_dimensions[r7].height = 22 if is_section else 18

    if is_section:
        ws7.merge_cells(f"A{r7}:F{r7}")
        bg = "1D4ED8" if "AWS" in label else ("059669" if "TWILIO" in label else ("D97706" if "FIREBASE" in label else ("0F172A" if "SCENARIO" in label else NAVY)))
        ws7.cell(r7,1).value = f"  {label}"
        ws7.cell(r7,1).font = fnt(10,True,WHITE)
        ws7.cell(r7,1).fill = F(bg)
        ws7.cell(r7,1).alignment = aln("left",indent=1)
        r7 += 1
        continue

    if is_total:
        row_bg = "0F172A"; tc7 = WHITE; sz7 = 10; bold7 = True
    elif is_warn:
        row_bg = "FEF3C7"; tc7 = "92400E"; sz7 = 9; bold7 = False
    else:
        row_bg = WHITE if r7 % 2 == 0 else BG_ALT; tc7 = NAVY; sz7 = 9; bold7 = False

    ws7.cell(r7,1).value = label
    ws7.cell(r7,1).font = fnt(sz7,bold7,tc7)
    ws7.cell(r7,1).fill = F(row_bg)
    ws7.cell(r7,1).alignment = aln(indent=1)

    for ci, val in [(2,s1l),(3,s1h),(4,s100l),(5,s100h)]:
        cl = ws7.cell(r7,ci)
        if val == 0 and not is_total:
            cl.value = "Free" if "Free" not in label and not is_warn else ("–" if is_total else "Free")
            # Determine display
            txt = label.lower()
            if "firebase" in txt or "fcm" in txt or "auth" in txt or "signalr" in txt or "route 53" in label:
                cl.value = "Free" if val == 0 else val
            else:
                cl.value = "–" if val == 0 else val
            cl.font = fnt(sz7, bold7, "065F46" if cl.value == "Free" else tc7)
        else:
            cl.value = val if val != 0 else ("Free" if not is_total else 0)
            cl.number_format = MONEY
            if is_total:
                cl.font = fnt(sz7,True,GOLD if ci >= 4 else "93C5FD")
            elif is_warn:
                cl.font = fnt(sz7,True,"991B1B" if ci >= 4 else "D97706")
            else:
                cl.font = fnt(sz7,bold7,tc7)
        cl.fill = F(row_bg)
        cl.alignment = aln("center")

    ws7.cell(r7,6).value = note
    ws7.cell(r7,6).font = fnt(8,False,MUTED if not is_total else "AAAAAA")
    ws7.cell(r7,6).fill = F(row_bg)
    ws7.cell(r7,6).alignment = aln(wrap=True)

    for c in range(1,7):
        ws7.cell(r7,c).border = btm_border()

    r7 += 1

# Recommendation box
r7 += 1
ws7.row_dimensions[r7].height = 32
ws7.merge_cells(f"A{r7}:F{r7}")
ws7[f"A{r7}"].value = (
    "✅  RECOMMENDATION: Use SignalR + PostgreSQL for member-to-member chat (already in your .NET 8 stack — zero extra cost). "
    "Reserve Twilio for VoIP calls + SMS OTP only. At 100K members this saves ₹42,000–₹84,000 per month (₹5–10 lakh/year). "
    "Twilio Conversations only makes sense if you need cross-platform chat history, typing indicators, and rich media outside the app context."
)
ws7[f"A{r7}"].font  = fnt(9,True,"065F46"); ws7[f"A{r7}"].fill = F("DCFCE7")
ws7[f"A{r7}"].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)

r7 += 2
ws7.row_dimensions[r7].height = 20
ws7.merge_cells(f"A{r7}:F{r7}")
ws7[f"A{r7}"].value = (
    "⚠️  AWS cost grows ~3.5× from 500-1K to 100K members due to: larger RDS instance class, more ECS tasks, significantly higher CloudFront egress, "
    "and two new mandatory services (ElastiCache + NAT Gateway). All costs are estimates; review quarterly and right-size as user base is confirmed."
)
ws7[f"A{r7}"].font  = fnt(9,False,NAVY); ws7[f"A{r7}"].fill = F("FEF3C7")
ws7[f"A{r7}"].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1)

# ── Save ──────────────────────────────────────────────────
out_path = "ANM_Resource_Cost_Model.xlsx"
wb.save(out_path)
print(f"\n✅  {out_path} generated successfully!")
print("   Open in Excel or LibreOffice Calc.")
print("\n   Key outputs (approximate — 10-week .NET 8 framework model):")
print("     Framework Advantage : saves ~4-6 weeks vs greenfield .NET 8 build")
print("     Core Duration       : 10 weeks + 30-day hypercare")
print("     Stack               : ASP.NET Core .NET 8 · React + Ionic/Capacitor · PostgreSQL 16 · AWS RDS")
print("     Total Days          : ~200-220 team-days across 12 weeks (incl. HC)")
print("\n   INR Pricing Packages:")
print("     Starter:    ₹32L  (10 weeks, core features)")
print("     Standard:   ₹38L  (10 weeks, full stack incl. Twilio + VoIP)")
print("     Enterprise: ₹48L  (10 weeks, push notifs + group chat + admin portal)")
print("\n   AMC Options (post go-live):")
print("     Basic:     ₹15,000/month")
print("     Standard:  ₹25,000/month")
print("     Premium:   ₹40,000/month")
print("\n   Monthly Platform Costs — 100K Members (client-borne):")
print("     AWS (RDS+ECS+S3+CF+ALB+CW+Redis+NAT+SES): ₹94,200–₹1,36,000/mo")
print("     Twilio Voice + SMS:                          ₹15,750–₹35,700/mo")
print("     Twilio Conversations (⚠️ if used):           ₹42,000–₹84,000/mo")
print("     Firebase Auth + FCM:                         Free–₹1,008/mo")
print("     ---")
print("     Scenario A (with Twilio Chat):               ₹1,52,454–₹2,56,708/mo")
print("     Scenario B (SignalR chat, recommended):      ₹1,10,454–₹1,72,708/mo")
print("     Annual saving using SignalR vs Twilio Chat:  ₹5.0L–₹10.1L/year")
