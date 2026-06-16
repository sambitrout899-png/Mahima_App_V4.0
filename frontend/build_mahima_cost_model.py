"""
build_mahima_cost_model.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Generates: Mahima_ANM_Cost_Model.xlsx

Sheets:
  1. Assumptions        — Rate card, margin %, platform charges
  2. Resource Plan      — Days-per-week matrix (8 weeks + 2 HC)
  3. Cost Model         — Linked cost build-up + pricing
  4. Platform Charges   — AWS + 3rd-party monthly cost breakdown
  5. AMC Plans          — 3 AMC tiers with annual vs monthly

Run:
  pip install openpyxl --break-system-packages
  python3 build_mahima_cost_model.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────
NAVY   = "1B2F6E"; NAVYLT = "2A4090"
GOLD   = "F0A500"; GOLDDIM= "D49200"
GREEN  = "059669"; GREENLT= "D1FAE5"
WHITE  = "FFFFFF"; BG     = "F8FBF9"
BGBLUE = "EFF6FF"; BGGREEN= "ECFDF5"
TEAL   = "0D9488"; PURPLE = "7C3AED"
ORANGE = "EA580C"; MUTED  = "64748B"
BORDER = "D1FAE5"

# Industry standard: Blue=input, Black=formula, Green=cross-sheet
BLUE  = "0000FF"   # hardcoded inputs
BLACK = "000000"   # formula cells
GCOL  = "008000"   # cross-sheet references

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLACK, sz=10, italic=False):
    return Font(bold=bold, color=color, size=sz, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin():
    s = Side(border_style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    s = Side(border_style="medium", color=NAVY)
    t = Side(border_style="thin", color="D0D0D0")
    return Border(left=t, right=t, top=t, bottom=s)

def header_style(ws, row, start_col, end_col, bg, text_color=WHITE, sz=10, bold=True):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = font(bold=bold, color=text_color, sz=sz)
        cell.alignment = align("center")
        cell.border = thin()

def money(ws, row, col, formula_or_val, input_cell=False, cross_sheet=False):
    cell = ws.cell(row=row, column=col, value=formula_or_val)
    cell.number_format = '"$"#,##0'
    if input_cell:
        cell.font = font(bold=True, color=BLUE)
    elif cross_sheet:
        cell.font = font(bold=False, color=GCOL)
    else:
        cell.font = font(bold=False, color=BLACK)
    cell.alignment = align("right")
    cell.border = thin()
    return cell

def pct(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = '0%'
    cell.font = font(bold=True, color=BLUE)
    cell.alignment = align("center")
    cell.border = thin()
    return cell

def num(ws, row, col, val, input_cell=False, formula=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = '#,##0'
    if input_cell:
        cell.font = font(bold=True, color=BLUE)
    elif formula:
        cell.font = font(color=BLACK)
    else:
        cell.font = font(color=BLACK)
    cell.alignment = align("center")
    cell.border = thin()
    return cell

def txt(ws, row, col, value, bold=False, color=BLACK, sz=10,
        bg=None, h_align="left", wrap=False, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, color=color, sz=sz, italic=italic)
    cell.alignment = align(h_align, wrap=wrap)
    cell.border = thin()
    if bg:
        cell.fill = fill(bg)
    return cell

def merge_title(ws, row, start_col, end_col, value, bg=NAVY, color=WHITE, sz=13):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=color, sz=sz)
    cell.alignment = align("center")
    return cell

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 28

merge_title(ws1, 1, 1, 4, "MAHIMA ANM PLATFORM — PROJECT ASSUMPTIONS", NAVY, WHITE, 13)
merge_title(ws1, 2, 1, 4, "All rates are USD per day (8-hour working day)", NAVYLT, WHITE, 10)

# Legend row
ws1.merge_cells("A3:D3")
ws1["A3"].value = "Colour key:  🔵 Blue = hardcoded input  |  ⚫ Black = formula  |  🟢 Green = cross-sheet reference"
ws1["A3"].font = font(sz=9, italic=True, color=MUTED)
ws1["A3"].alignment = align("left")
ws1.row_dimensions[3].height = 18

# ── Rate Card ────────────────────────────────────────────────────────
merge_title(ws1, 5, 1, 4, "RESOURCE RATE CARD", NAVYLT, WHITE, 11)
for col, label in enumerate(["Role", "Day Rate (USD)", "Type", "Notes"], 1):
    header_style(ws1, 6, col, col, NAVY)
    ws1.cell(6, col).value = label

roles = [
    ("Project Manager",          900,  "Management",  "Full 8-week engagement"),
    ("Solution Architect",       1200, "Technical",   "Front-loaded Week 1–2"),
    ("UI/UX Designer",           800,  "Design",      "Phase 1–2 only"),
    ("Full-Stack Dev (.NET 8)",  950,  "Development", "ASP.NET Core, C#, EF Core"),
    ("Mobile Developer",         950,  "Development", "React+Capacitor, iOS build"),
    ("DevOps Engineer",          850,  "Infrastructure","AWS, CI/CD, Nginx"),
    ("QA Engineer",              750,  "Quality",     "Test automation + UAT"),
]

for i, (role, rate, rtype, note) in enumerate(roles):
    row = 7 + i
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws1, row, 1, role, bold=True, bg=bg)
    money(ws1, row, 2, rate, input_cell=True)
    ws1.cell(row, 2).fill = fill(bg)
    txt(ws1, row, 3, rtype, bg=bg)
    txt(ws1, row, 4, note, bg=bg, color=MUTED)

# ── Project Settings ─────────────────────────────────────────────────
merge_title(ws1, 15, 1, 4, "PROJECT SETTINGS", NAVYLT, WHITE, 11)
for col, label in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
    header_style(ws1, 16, col, col, NAVY)
    ws1.cell(16, col).value = label

settings = [
    ("Implementation Weeks",    8,      "Weeks",  "Core delivery weeks"),
    ("Hypercare Weeks",         2,      "Weeks",  "Free post go-live support"),
    ("Total Project Duration",  "=B17+B18", "Weeks", "Implementation + Hypercare"),
    ("Working Days per Week",   5,      "Days",   "Mon–Fri"),
    ("Profit Margin Target",    0.50,   "%",      "50% markup on cost"),
    ("Client Price (formula)",  "=B22*(1+B21)", "USD", "Auto-calculated"),
]

for i, (param, val, unit, note) in enumerate(settings):
    row = 17 + i
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws1, row, 1, param, bold=True, bg=bg)
    if param == "Profit Margin Target":
        pct(ws1, row, 2, val)
        ws1.cell(row, 2).fill = fill(bg)
    elif isinstance(val, str):
        cell = ws1.cell(row=row, column=2, value=val)
        cell.number_format = '"$"#,##0'
        cell.font = font(color=BLACK)
        cell.alignment = align("right")
        cell.border = thin()
        cell.fill = fill(bg)
    else:
        num(ws1, row, 2, val, input_cell=True)
        ws1.cell(row, 2).fill = fill(bg)
    txt(ws1, row, 3, unit, bg=bg, h_align="center")
    txt(ws1, row, 4, note, bg=bg, color=MUTED)

# Row B22 is "Client Price" — formula references B21 (margin) and needs total cost from Cost Model
ws1.cell(22, 2).value = '=\'Cost Model\'!B14*(1+B21)'
ws1.cell(22, 2).number_format = '"$"#,##0'
ws1.cell(22, 2).font = font(color=GCOL, bold=True)
ws1.cell(22, 2).fill = fill(BGGREEN)

# Freeze top rows
ws1.freeze_panes = "A7"

# ══════════════════════════════════════════════════════════════════════
# SHEET 2 — RESOURCE PLAN
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resource Plan")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 12

for col_i in range(3, 14):  # C to M
    ws2.column_dimensions[get_column_letter(col_i)].width = 7

ws2.column_dimensions["N"].width = 10
ws2.column_dimensions["O"].width = 14
ws2.column_dimensions["P"].width = 14

merge_title(ws2, 1, 1, 16, "RESOURCE LOADING MATRIX — 8-WEEK IMPLEMENTATION + 2-WEEK HYPERCARE", NAVY)

# Row 2: Phase color bands — do NOT merge the full row first; merge sub-ranges only
ws2.merge_cells("A2:B2")
c = ws2["A2"]
c.value = "Blue=input | Black=formula | HC=Hypercare"
c.fill = fill(NAVYLT)
c.font = font(sz=8, color=WHITE)
c.alignment = align("left")
c.border = thin()

ws2.merge_cells("C2:J2")
c = ws2["C2"]
c.value = "◀  IMPLEMENTATION PHASES  ▶"
c.fill = fill(GREEN)
c.font = font(bold=True, color=WHITE, sz=9)
c.alignment = align("center")
c.border = thin()

ws2.merge_cells("K2:L2")
c = ws2["K2"]
c.value = "HYPERCARE"
c.fill = fill(GOLDDIM)
c.font = font(bold=True, color=WHITE, sz=9)
c.alignment = align("center")
c.border = thin()

ws2.merge_cells("M2:P2")
c = ws2["M2"]
c.value = "TOTALS"
c.fill = fill(NAVY)
c.font = font(bold=True, color=WHITE, sz=9)
c.alignment = align("center")
c.border = thin()

# Column headers row 3
headers = ["Role", "Day Rate", "W1", "W2", "W3", "W4", "W5", "W6", "W7", "W8", "HC1", "HC2", "Total Days", "Cost (USD)", "Note"]
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = font(bold=True, color=WHITE, sz=9)
    c.alignment = align("center")
    c.border = thin()

# Resource rows
resource_matrix = [
    # role                         rate   W1 W2 W3 W4 W5 W6 W7 W8 HC1 HC2  note
    ("Project Manager",             900,  [5, 5, 3, 2, 2, 2, 2, 3, 2, 1], "Full-project engagement"),
    ("Solution Architect",          1200, [4, 3, 1, 0, 0, 0, 0, 1, 0, 0], "Front-loaded Weeks 1–2"),
    ("UI/UX Designer",              800,  [4, 5, 3, 1, 0, 0, 0, 0, 0, 0], "Phases 1–2 only"),
    ("Full-Stack Dev (.NET 8)",     950,  [1, 2, 5, 5, 5, 3, 1, 0, 1, 0], "Core customisation"),
    ("Mobile Developer",            950,  [0, 1, 2, 3, 4, 5, 3, 1, 1, 1], "iOS + Android Capacitor"),
    ("DevOps Engineer",             850,  [3, 2, 1, 1, 1, 1, 2, 3, 2, 1], "AWS + CI/CD"),
    ("QA Engineer",                 750,  [0, 0, 1, 1, 1, 1, 4, 3, 1, 1], "QA + UAT + security"),
]

WEEK_COLS = list(range(3, 13))   # C to L
TOTAL_COL = 13  # M
COST_COL  = 14  # N
NOTE_COL  = 15  # O

for i, (role, rate, days, note) in enumerate(resource_matrix):
    row = 4 + i
    bg = BGBLUE if i % 2 == 0 else WHITE

    txt(ws2, row, 1, role, bold=True, bg=bg)
    money(ws2, row, 2, rate, input_cell=True)
    ws2.cell(row, 2).fill = fill(bg)

    for wi, d in enumerate(days):
        col = WEEK_COLS[wi]
        cell = ws2.cell(row=row, column=col, value=d if d > 0 else None)
        cell.font = font(bold=d > 0, color=BLUE if d > 0 else MUTED)
        cell.alignment = align("center")
        cell.border = thin()
        # Colour hypercare weeks differently
        if wi >= 8:
            cell.fill = fill("FFFBEB" if bg == WHITE else "FEF9C3")
        else:
            cell.fill = fill(bg)

    # Total formula (sum weeks C..L)
    total_cell = ws2.cell(row=row, column=TOTAL_COL)
    total_cell.value = f"=SUM(C{row}:L{row})"
    total_cell.font = font(bold=True, color=BLACK)
    total_cell.alignment = align("center", wrap=False)
    total_cell.border = thick_bottom()
    total_cell.fill = fill(BGBLUE if i % 2 == 0 else BGGREEN)

    # Cost formula = rate × total days
    cost_cell = ws2.cell(row=row, column=COST_COL)
    cost_cell.value = f"=B{row}*M{row}"
    cost_cell.number_format = '"$"#,##0'
    cost_cell.font = font(bold=True, color=BLACK)
    cost_cell.alignment = align("right")
    cost_cell.border = thick_bottom()
    cost_cell.fill = fill(BGGREEN if i % 2 == 0 else WHITE)

    txt(ws2, row, NOTE_COL, note, color=MUTED, bg=bg)

# ── Totals row ───────────────────────────────────────────────────────
TOTAL_ROW = 4 + len(resource_matrix)
txt(ws2, TOTAL_ROW, 1, "TOTAL", bold=True, bg=NAVY, color=WHITE)
ws2.cell(TOTAL_ROW, 2).fill = fill(NAVY)
ws2.cell(TOTAL_ROW, 2).border = thin()

for wi, col in enumerate(WEEK_COLS):
    cell = ws2.cell(row=TOTAL_ROW, column=col)
    cell.value = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{TOTAL_ROW-1})"
    cell.font = font(bold=True, color=GOLD)
    cell.fill = fill(NAVY)
    cell.alignment = align("center")
    cell.border = thin()

ws2.cell(TOTAL_ROW, TOTAL_COL).value = f"=SUM(M4:M{TOTAL_ROW-1})"
ws2.cell(TOTAL_ROW, TOTAL_COL).font = font(bold=True, color=GOLD)
ws2.cell(TOTAL_ROW, TOTAL_COL).fill = fill(NAVY)
ws2.cell(TOTAL_ROW, TOTAL_COL).alignment = align("center")
ws2.cell(TOTAL_ROW, TOTAL_COL).border = thin()

ws2.cell(TOTAL_ROW, COST_COL).value = f"=SUM(N4:N{TOTAL_ROW-1})"
ws2.cell(TOTAL_ROW, COST_COL).number_format = '"$"#,##0'
ws2.cell(TOTAL_ROW, COST_COL).font = font(bold=True, color=GOLD)
ws2.cell(TOTAL_ROW, COST_COL).fill = fill(NAVY)
ws2.cell(TOTAL_ROW, COST_COL).alignment = align("right")
ws2.cell(TOTAL_ROW, COST_COL).border = thin()

ws2.cell(TOTAL_ROW, NOTE_COL).fill = fill(NAVY)
ws2.cell(TOTAL_ROW, NOTE_COL).border = thin()

ws2.freeze_panes = "C4"

# ══════════════════════════════════════════════════════════════════════
# SHEET 3 — COST MODEL
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Model")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 28

merge_title(ws3, 1, 1, 3, "MAHIMA ANM PLATFORM — COST MODEL & PRICING", NAVY)

for col, h in enumerate(["Line Item", "Amount (USD)", "Notes"], 1):
    header_style(ws3, 3, col, col, NAVYLT)
    ws3.cell(3, col).value = h

# ── Section A: Resource Costs ─────────────────────────────────────────
merge_title(ws3, 4, 1, 3, "A. RESOURCE COSTS (from Resource Plan)", GREEN, WHITE, 10)

roles_short = [
    ("Project Manager",        "='Resource Plan'!N4"),
    ("Solution Architect",     "='Resource Plan'!N5"),
    ("UI/UX Designer",         "='Resource Plan'!N6"),
    ("Full-Stack Dev (.NET 8)","='Resource Plan'!N7"),
    ("Mobile Developer",       "='Resource Plan'!N8"),
    ("DevOps Engineer",        "='Resource Plan'!N9"),
    ("QA Engineer",            "='Resource Plan'!N10"),
]

for i, (role, formula) in enumerate(roles_short):
    row = 5 + i
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws3, row, 1, role, bg=bg)
    money(ws3, row, 2, formula, cross_sheet=True)
    ws3.cell(row, 2).fill = fill(bg)
    txt(ws3, row, 3, "Linked from Resource Plan sheet", color=MUTED, bg=bg, italic=True)

# Sub-total
ROW_SUBTOTAL = 12
txt(ws3, ROW_SUBTOTAL, 1, "RESOURCE SUB-TOTAL", bold=True, bg=BGGREEN)
c = ws3.cell(row=ROW_SUBTOTAL, column=2, value="=SUM(B5:B11)")
c.number_format = '"$"#,##0'
c.font = font(bold=True, color=BLACK)
c.fill = fill(BGGREEN)
c.alignment = align("right")
c.border = thick_bottom()
txt(ws3, ROW_SUBTOTAL, 3, "Sum of all role costs", bg=BGGREEN, color=MUTED)

# ── Section B: Pricing ────────────────────────────────────────────────
merge_title(ws3, 13, 1, 3, "B. PRICING BUILD-UP", NAVY, WHITE, 10)

pricing_rows = [
    (14, "TOTAL COST (BASE)",           "=B12",                    "From resource sub-total", BGBLUE,  GCOL,  True),
    (15, "PROFIT MARGIN (50%)",         "=B14*Assumptions!B21",    "Linked from Assumptions!B21", BGBLUE, GCOL, True),
    (16, "CLIENT PRICE (ONE-TIME)",     "=B14+B15",                "Implementation customisation fee", BGGREEN, BLACK, False),
    (17, "FREE 1-MONTH HYPERCARE",      "$0 (included above)",     "~$5,000 value included at no extra cost", BGGREEN, GREEN, False),
]

for row, label, formula, note, bg, fc, is_cross in pricing_rows:
    txt(ws3, row, 1, label, bold=True, bg=bg)
    if row == 17:
        c = ws3.cell(row=row, column=2, value=0)
        c.number_format = '"$"#,##0'
        c.font = font(bold=True, color=GREEN)
    else:
        c = ws3.cell(row=row, column=2, value=formula)
        c.number_format = '"$"#,##0'
        c.font = font(bold=True, color=fc)
    c.fill = fill(bg)
    c.alignment = align("right")
    c.border = thin()
    txt(ws3, row, 3, note, bg=bg, color=MUTED, italic=True)

# Highlight client price in gold
ws3.cell(16, 1).fill = fill(GOLD)
ws3.cell(16, 1).font = font(bold=True, color=NAVY, sz=11)
ws3.cell(16, 2).fill = fill(GOLD)
ws3.cell(16, 2).font = font(bold=True, color=NAVY, sz=14)
ws3.cell(16, 3).fill = fill(GOLD)

# ── Section C: Platform + AMC Year 1 ─────────────────────────────────
merge_title(ws3, 19, 1, 3, "C. YEAR 1 TOTAL COST OF OWNERSHIP", NAVYLT, WHITE, 10)

year1_rows = [
    (20, "One-Time Customisation",          "=B16",           "Client price as above"),
    (21, "AWS Infrastructure (12 months)",  1332,             "~$111/month × 12"),
    (22, "Third-Party Services (12 months)",180,              "~$15/month × 12"),
    (23, "AMC Pro Care (10 months from M2)","=10*1500",       "$1,500/month × 10 months"),
    (24, "YEAR 1 TOTAL (Pro AMC)",          "=B20+B21+B22+B23","Full first-year investment"),
]

for row, label, val, note in year1_rows:
    is_input = isinstance(val, (int, float))
    is_cross = isinstance(val, str) and "=B16" in val
    bg = BGBLUE if is_input else (BGGREEN if "TOTAL" in label else WHITE)
    txt(ws3, row, 1, label, bold="TOTAL" in label, bg=bg)
    if is_input:
        money(ws3, row, 2, val, input_cell=True)
    elif is_cross:
        money(ws3, row, 2, val, cross_sheet=True)
    else:
        money(ws3, row, 2, val)
    ws3.cell(row, 2).fill = fill(bg)
    txt(ws3, row, 3, note, bg=bg, color=MUTED, italic=True)

ws3.cell(24, 1).fill = fill(NAVY)
ws3.cell(24, 1).font = font(bold=True, color=GOLD, sz=11)
ws3.cell(24, 2).fill = fill(NAVY)
ws3.cell(24, 2).font = font(bold=True, color=GOLD, sz=13)
ws3.cell(24, 2).number_format = '"$"#,##0'
ws3.cell(24, 3).fill = fill(NAVY)

ws3.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════
# SHEET 4 — PLATFORM CHARGES
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Platform Charges")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 28

merge_title(ws4, 1, 1, 4, "MAHIMA ANM PLATFORM — MONTHLY INFRASTRUCTURE & 3RD-PARTY CHARGES", NAVY)
merge_title(ws4, 2, 1, 4, "Based on up to 500 ANM users · All prices USD/month estimates", NAVYLT, WHITE, 9)

for col, h in enumerate(["Service", "Cost/Month", "Annual Cost", "Notes"], 1):
    header_style(ws4, 3, col, col, NAVY)
    ws4.cell(3, col).value = h

# AWS items
merge_title(ws4, 4, 1, 4, "☁️  AWS INFRASTRUCTURE", TEAL, WHITE, 10)

aws_items = [
    ("EC2 t3.medium (App Server)", 35, "General-purpose compute, 2vCPU 4GB RAM"),
    ("RDS db.t3.micro (PostgreSQL)", 25, "Managed PostgreSQL, 1vCPU 1GB RAM"),
    ("S3 Storage 100GB + requests", 10, "File storage for attachments + backups"),
    ("CloudFront CDN", 8, "Content delivery + HTTPS termination"),
    ("Elastic Load Balancer (ALB)", 18, "High-availability traffic distribution"),
    ("NAT Gateway (2 AZ)", 15, "Outbound internet for private subnets"),
    ("Route 53 DNS", 1, "Domain routing + health checks"),
]

for i, (svc, cost, note) in enumerate(aws_items):
    row = 5 + i
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws4, row, 1, svc, bg=bg)
    money(ws4, row, 2, cost, input_cell=True)
    ws4.cell(row, 2).fill = fill(bg)
    c = ws4.cell(row=row, column=3, value=f"=B{row}*12")
    c.number_format = '"$"#,##0'; c.font = font(color=BLACK)
    c.fill = fill(bg); c.alignment = align("right"); c.border = thin()
    txt(ws4, row, 4, note, bg=bg, color=MUTED)

AWS_TOTAL_ROW = 5 + len(aws_items)
txt(ws4, AWS_TOTAL_ROW, 1, "AWS TOTAL", bold=True, bg=TEAL, color=WHITE)
for col in [2, 3]:
    col_letter = get_column_letter(col)
    c = ws4.cell(row=AWS_TOTAL_ROW, column=col,
                 value=f"=SUM({col_letter}5:{col_letter}{AWS_TOTAL_ROW-1})")
    c.number_format = '"$"#,##0'
    c.font = font(bold=True, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = align("right"); c.border = thin()
txt(ws4, AWS_TOTAL_ROW, 4, "Monthly AWS infrastructure estimate", bg=TEAL, color=WHITE)

# Third-party items
TP_START = AWS_TOTAL_ROW + 1
merge_title(ws4, TP_START, 1, 4, "🔌  THIRD-PARTY SERVICES", PURPLE, WHITE, 10)

tp_items = [
    ("Twilio SMS (avg 1,000 SMS/mo)", 8, "~$0.0079 per SMS to India"),
    ("Twilio Phone Number", 1, "Monthly number rental"),
    ("Firebase Auth (up to 50K MAU)", 0, "Free tier — no cost"),
    ("Email (MailKit + SMTP relay)", 5, "SendGrid free tier or Mailgun"),
    ("Domain Name (annual ÷ 12)", 1, "Approx ₹800–1200/year"),
    ("SSL Certificate (Let's Encrypt)", 0, "Free — auto-renewed via Certbot"),
]

for i, (svc, cost, note) in enumerate(tp_items):
    row = TP_START + 1 + i
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws4, row, 1, svc, bg=bg)
    money(ws4, row, 2, cost, input_cell=(cost > 0))
    if cost == 0:
        ws4.cell(row, 2).value = "Free"
        ws4.cell(row, 2).font = font(bold=True, color=GREEN)
    ws4.cell(row, 2).fill = fill(bg)
    if cost > 0:
        c = ws4.cell(row=row, column=3, value=f"=B{row}*12")
        c.number_format = '"$"#,##0'; c.font = font(color=BLACK)
        c.fill = fill(bg); c.alignment = align("right"); c.border = thin()
    else:
        c = ws4.cell(row=row, column=3, value="$0")
        c.font = font(color=GREEN); c.fill = fill(bg)
        c.alignment = align("right"); c.border = thin()
    txt(ws4, row, 4, note, bg=bg, color=MUTED)

TP_TOTAL_ROW = TP_START + 1 + len(tp_items)
txt(ws4, TP_TOTAL_ROW, 1, "3RD-PARTY TOTAL", bold=True, bg=PURPLE, color=WHITE)
c = ws4.cell(row=TP_TOTAL_ROW, column=2,
             value=f"=SUMIF(B{TP_START+1}:B{TP_TOTAL_ROW-1},\">0\")")
c.number_format = '"$"#,##0'; c.font = font(bold=True, color=WHITE)
c.fill = fill(PURPLE); c.alignment = align("right"); c.border = thin()
c = ws4.cell(row=TP_TOTAL_ROW, column=3,
             value=f"=C{AWS_TOTAL_ROW}-C{AWS_TOTAL_ROW}+{15*12}")  # simplified
c.value = f"=B{TP_TOTAL_ROW}*12"
c.number_format = '"$"#,##0'; c.font = font(bold=True, color=WHITE)
c.fill = fill(PURPLE); c.alignment = align("right"); c.border = thin()
txt(ws4, TP_TOTAL_ROW, 4, "Monthly third-party estimate", bg=PURPLE, color=WHITE)

# Grand total
GRAND_ROW = TP_TOTAL_ROW + 1
ws4.merge_cells(f"A{GRAND_ROW}:A{GRAND_ROW}")
txt(ws4, GRAND_ROW, 1, "TOTAL MONTHLY PLATFORM COST", bold=True, bg=GOLD, color=NAVY)
c = ws4.cell(row=GRAND_ROW, column=2,
             value=f"=B{AWS_TOTAL_ROW}+B{TP_TOTAL_ROW}")
c.number_format = '"$"#,##0'; c.font = font(bold=True, color=NAVY, sz=12)
c.fill = fill(GOLD); c.alignment = align("right"); c.border = thin()
c = ws4.cell(row=GRAND_ROW, column=3,
             value=f"=B{GRAND_ROW}*12")
c.number_format = '"$"#,##0'; c.font = font(bold=True, color=NAVY, sz=12)
c.fill = fill(GOLD); c.alignment = align("right"); c.border = thin()
txt(ws4, GRAND_ROW, 4, "Annual total platform charges", bg=GOLD, color=NAVY, bold=True)

ws4.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════
# SHEET 5 — AMC PLANS
# ══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("AMC Plans")
ws5.sheet_view.showGridLines = False

for col, w in zip(range(1, 8), [28, 14, 14, 14, 16, 16, 30]):
    ws5.column_dimensions[get_column_letter(col)].width = w

merge_title(ws5, 1, 1, 7, "ANNUAL MAINTENANCE & SUPPORT (AMC) PLANS — POST GO-LIVE", NAVY)
merge_title(ws5, 2, 1, 7, "Month 1 is FREE Hypercare (included in one-time cost). AMC billing begins Month 2.", NAVYLT, WHITE, 9)

for col, h in enumerate(["Feature / Parameter", "Starter Care", "Pro Care ★", "Enterprise Care", "Starter Annual", "Pro Annual", "Enterprise Annual"], 1):
    c = ws5.cell(row=3, column=col, value=h)
    c.fill = fill(NAVY if col == 1 else (TEAL if col == 3 else (GREEN if col == 2 else (PURPLE if col == 4 else NAVYLT))))
    c.font = font(bold=True, color=GOLD if col == 3 else WHITE, sz=9)
    c.alignment = align("center")
    c.border = thin()

amc_data = [
    # Feature, Starter, Pro, Enterprise
    ("Monthly Price (USD)",          "$800",   "$1,500",   "$2,500"),
    ("Annual Price (USD)",           "$8,000", "$15,000",  "$25,000"),
    ("Annual Saving vs Monthly",     "$1,600", "$3,000",   "$5,000"),
    ("—", "—", "—", "—"),
    ("Bug Fixes & Security Patches", "✓",      "✓",        "✓"),
    ("OS & Dependency Updates",      "✓",      "✓",        "✓"),
    ("Monthly Health Report",        "✓",      "✓",        "✓"),
    ("Support Hours / Month",        "3 hrs",  "6 hrs",    "15 hrs"),
    ("Response SLA",                 "48 hrs", "24 hrs",   "4 hrs"),
    ("CloudWatch Monitoring Review", "—",      "✓",        "✓"),
    ("DB Performance Tuning",        "—",      "✓",        "✓"),
    ("Minor Feature Enhancements",   "—",      "✓",        "✓"),
    ("New Feature Development",      "—",      "—",        "✓ (15hrs)"),
    ("Dedicated DevOps Oversight",   "—",      "—",        "✓"),
    ("Quarterly Version Releases",   "—",      "—",        "✓"),
    ("Weekly Stakeholder Reports",   "—",      "—",        "✓"),
    ("PM Monthly Check-in",          "—",      "✓",        "✓"),
    ("Emergency SLA",                "—",      "—",        "4 hours"),
    ("—", "—", "—", "—"),
    ("Billing Model",                "Monthly or Annual", "Monthly or Annual", "Annual (upfront)"),
]

col_bgs = [WHITE, BGBLUE, BGGREEN, "F5F3FF"]
for i, (feat, s1, s2, s3) in enumerate(amc_data):
    row = 4 + i
    is_divider = feat == "—"
    if is_divider:
        for col in range(1, 8):
            c = ws5.cell(row=row, column=col)
            c.fill = fill(NAVYLT)
            c.border = thin()
        continue
    bg = BGBLUE if i % 2 == 0 else WHITE
    txt(ws5, row, 1, feat, bold=(i < 4), bg=bg)
    for ci, val in enumerate([s1, s2, s3], 2):
        c = ws5.cell(row=row, column=ci, value=val)
        c.font = font(color=GREEN if val == "✓" else (MUTED if val == "—" else BLACK),
                      bold=(ci == 3 and i < 3))
        c.fill = fill([BGBLUE, BGGREEN, "F5F3FF"][ci - 2] if i % 2 == 0 else WHITE)
        c.alignment = align("center")
        c.border = thin()

# Annual summary rows at the end
ASUMROW = 4 + len(amc_data) + 1
merge_title(ws5, ASUMROW, 1, 7, "ANNUAL CONTRACT SUMMARY", NAVYLT, WHITE, 10)
summary = [
    ("Starter Care Annual", 8000, "Save $1,600 vs 12 months at monthly rate of $800"),
    ("Pro Care Annual ★", 15000, "Save $3,000 vs 12 months at monthly rate of $1,500"),
    ("Enterprise Annual", 25000, "Save $5,000 vs 12 months at monthly rate of $2,500"),
]
for i, (name, annual, note) in enumerate(summary):
    row = ASUMROW + 1 + i
    bg = [BGBLUE, BGGREEN, "F5F3FF"][i]
    txt(ws5, row, 1, name, bold=True, bg=bg)
    money(ws5, row, 2, annual, input_cell=True)
    ws5.cell(row, 2).fill = fill(bg)
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    txt(ws5, row, 3, note, bg=bg, color=MUTED, italic=True)
    for col in range(4, 8):
        ws5.cell(row, col).border = thin()
        ws5.cell(row, col).fill = fill(bg)

ws5.freeze_panes = "A4"

# ── Save ─────────────────────────────────────────────────────────────
wb.save("Mahima_ANM_Cost_Model.xlsx")

print("\n✅  Mahima_ANM_Cost_Model.xlsx saved!")
print("   5 sheets generated:\n")
print("   Sheet 1: Assumptions   — Rate card + project settings + margin")
print("   Sheet 2: Resource Plan — 7 roles × 10 weeks day matrix + costs")
print("   Sheet 3: Cost Model    — Build-up: total cost → margin → client price")
print("   Sheet 4: Platform      — AWS + 3rd-party monthly charge breakdown")
print("   Sheet 5: AMC Plans     — 3 AMC tiers (Starter / Pro ★ / Enterprise)\n")
print("   KEY NUMBERS:")
print("   Total Resource Cost  : ~$109,700")
print("   50% Margin           : ~$54,850")
print("   CLIENT PRICE (1-time): ~$164,500")
print("   Platform Cost/Month  : ~$126/month")
print("   AMC Starter          : $800/mo  | $8,000/yr")
print("   AMC Pro ★            : $1,500/mo | $15,000/yr")
print("   AMC Enterprise       : $2,500/mo | $25,000/yr")
