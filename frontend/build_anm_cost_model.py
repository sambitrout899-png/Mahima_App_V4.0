"""
ANM Community App — Resource Loading & Cost Model
══════════════════════════════════════════════════
Run:
    pip install openpyxl
    python build_anm_cost_model.py

Output: ANM_Resource_Cost_Model.xlsx  (4 sheets)

Sheets
  1. Assumptions  — rate card, margin %, project parameters
  2. Resource Plan — week-by-week days per role (12 weeks incl. hypercare)
  3. Cost Model   — cost build-up, 50% margin, client price, phase breakdown
  4. Weekly Cost  — weekly & cumulative cost + selling price
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Palette ───────────────────────────────────────────────
NAVY   = "1B2F6E"
GOLD   = "F0A500"
WHITE  = "FFFFFF"
MUTED  = "64748B"
BLUE   = "0000FF"   # hardcoded inputs
BLACK  = "000000"   # formula cells
GREEN  = "008000"   # cross-sheet links
BG_ALT = "F5F7FA"
BG_HDR = "EEF2F7"

# ── Helpers ───────────────────────────────────────────────
def F(c): return PatternFill("solid", fgColor=c)
def fnt(sz=10, bold=False, color=BLACK, name="Arial"): return Font(name=name, size=sz, bold=bold, color=color)
def aln(h="left", v="center", wrap=False, indent=0): return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def btm_border(): return Border(bottom=Side(style="thin", color="DDDDDD"))

def hdr_cell(ws, r, c, val, bg=NAVY, tc=WHITE, sz=10, bold=True, h="left", indent=1):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(sz, bold, tc)
    cl.fill = F(bg); cl.alignment = aln(h, indent=indent)

def inp(ws, r, c, val, fmt=None, bg=None, h="center"):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(color=BLUE)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

def calc(ws, r, c, val, fmt=None, tc=BLACK, bg=None, h="right", sz=10, bold=False):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(sz, bold, tc)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

def xlink(ws, r, c, val, fmt=None, h="center", bg=None):
    cl = ws.cell(r, c); cl.value = val; cl.font = fnt(color=GREEN)
    cl.alignment = aln(h)
    if fmt: cl.number_format = fmt
    if bg:  cl.fill = F(bg)

MONEY = "$#,##0;($#,##0);-"
PCT   = "0.0%;(0.0%);-"

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS
# Rate card rows 13-20  (referenced by Cost Model & Weekly Cost)
# Pricing row 25 = Margin %  (referenced by Cost Model & Weekly Cost)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"
ws1.column_dimensions["A"].width = 34
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 46

# Title
ws1.row_dimensions[1].height = 44
ws1.merge_cells("A1:D1")
ws1["A1"].value = "ANM Community App — Resource & Pricing Assumptions"
ws1["A1"].font  = fnt(18, True, NAVY)
ws1["A1"].fill  = F(BG_HDR)
ws1["A1"].alignment = aln("left", indent=1)

# Metadata rows 3-8
meta = [("Project","ANM Community App"),("Client","ANM — Church & Community Platform"),
        ("Prepared by","Solution Development Team"),("Date","June 2025"),
        ("Currency","USD ($)"),("Version","v1.0 — Draft")]
for i, (k, v) in enumerate(meta):
    r = 3 + i
    ws1.row_dimensions[r].height = 17
    ws1.cell(r,1).value = k; ws1.cell(r,1).font = fnt(9, True, MUTED)
    ws1.cell(r,2).value = v; ws1.cell(r,2).font = fnt(10, False, "1A1A2E")

# Gap
ws1.row_dimensions[10].height = 8

# ── Rate Card (rows 11-20) ────────────────────────────────
ws1.row_dimensions[11].height = 22
ws1.merge_cells("A11:D11")
hdr_cell(ws1, 11, 1, "  RATE CARD — Daily Rates (USD)")

ws1.row_dimensions[12].height = 18
for c, h in [(1,"Role"),(2,"Daily Rate (USD)"),(3,"Role ID"),(4,"Notes / Scope")]:
    ws1.cell(12,c).value = h; ws1.cell(12,c).font = fnt(9,True,NAVY)
    ws1.cell(12,c).fill  = F(BG_HDR); ws1.cell(12,c).alignment = aln("center")
    ws1.cell(12,c).border = Border(bottom=Side(style="medium", color=NAVY))

# Rows 13-20 = 8 roles  ← CRITICAL: Cost Model links to B13:B20
rates = [
    ("Project Manager",          900,  "PM",  "Sr PM; reporting, risk, ANM stakeholder comms"),
    ("Solution Architect",      1200,  "ARC", "Architecture reviews & formal sign-off gates"),
    ("UI/UX Designer",           800,  "DES", "Figma, design system, UAT design support"),
    ("Mobile Developer (Sr.)",   950,  "MOB", "React Native — iOS & Android cross-platform"),
    ("Backend Developer (Sr.)",  950,  "BAK", "Node.js / Express, MongoDB Atlas, REST APIs"),
    ("DevOps Engineer",          850,  "DEV", "AWS ECS, CI/CD, S3/CloudFront, monitoring"),
    ("QA Engineer",              750,  "QA",  "E2E, unit, load, accessibility testing"),
    ("Security Consultant",     1100,  "SEC", "OWASP pentest, audit & sign-off report"),
]
for i, (rname, rate, rid, note) in enumerate(rates):
    r = 13 + i
    ws1.row_dimensions[r].height = 18
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = rname; ws1.cell(r,1).font = fnt(10,True,NAVY);  ws1.cell(r,1).alignment = aln()
    inp(ws1, r, 2, rate, "$#,##0", bg)
    ws1.cell(r,3).value = rid;   ws1.cell(r,3).font = fnt(9,False,MUTED); ws1.cell(r,3).alignment = aln("center")
    ws1.cell(r,4).value = note;  ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).alignment = aln()
    for c in range(1,5):
        ws1.cell(r,c).fill   = F(bg)
        ws1.cell(r,c).border = btm_border()

# Gap
ws1.row_dimensions[21].height = 8
ws1.row_dimensions[22].height = 8

# ── Pricing Assumptions (rows 23-31) ─────────────────────
ws1.row_dimensions[23].height = 22
ws1.merge_cells("A23:D23")
hdr_cell(ws1, 23, 1, "  PRICING ASSUMPTIONS")

ws1.row_dimensions[24].height = 18
for c, h in [(1,"Parameter"),(2,"Value"),(4,"Notes")]:
    ws1.cell(24,c).value = h; ws1.cell(24,c).font = fnt(9,True,NAVY)
    ws1.cell(24,c).fill  = F(BG_HDR); ws1.cell(24,c).alignment = aln("center")
    ws1.cell(24,c).border = Border(bottom=Side(style="medium", color=NAVY))

# Row 25 = Margin %  ← referenced as Assumptions!B25 throughout
pricing_p = [
    ("Gross Margin % (on cost)",      0.50, "0.0%",  "50% margin added on top of total cost → client price"),
    ("Working Days / Week",           5,    "#,##0",  "Mon–Fri; excludes public holidays & bank holidays"),
    ("Hypercare Duration (weeks)",    2,    "#,##0",  "Post go-live monitoring & support period"),
    ("Sprint Length (weeks)",         2,    "#,##0",  "Agile: 2-week sprint cadence (5 sprints total)"),
    ("Project Duration — Core (wks)", 10,   "#,##0",  "Wk 1 through Wk 10 (Discovery → Go-Live)"),
    ("Total Duration incl. HC (wks)", 12,   "#,##0",  "Core 10 weeks + 2 hypercare weeks"),
]
for i, (param, val, fmt, note) in enumerate(pricing_p):
    r = 25 + i
    ws1.row_dimensions[r].height = 18
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = param; ws1.cell(r,1).font = fnt(10,True,NAVY);  ws1.cell(r,1).alignment = aln()
    inp(ws1, r, 2, val, fmt, bg)
    ws1.cell(r,4).value = note;  ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).alignment = aln()
    for c in [1,2,4]:
        ws1.cell(r,c).fill   = F(bg)
        ws1.cell(r,c).border = btm_border()

# Gap
ws1.row_dimensions[32].height = 8

# ── Color legend ──────────────────────────────────────────
ws1.row_dimensions[33].height = 22
ws1.merge_cells("A33:D33")
hdr_cell(ws1, 33, 1, "  COLOR CODING LEGEND")
legend_items = [
    ("Blue text",  BLUE,  "Hardcoded input — users may change these for scenario analysis"),
    ("Black text", BLACK, "Excel formula / calculated value — do not edit manually"),
    ("Green text", GREEN, "Link pulling from another worksheet within this workbook"),
]
for i, (lbl_, tc, note) in enumerate(legend_items):
    r = 34 + i
    ws1.row_dimensions[r].height = 17
    bg = WHITE if i%2 == 0 else BG_ALT
    ws1.cell(r,1).value = lbl_;       ws1.cell(r,1).font = fnt(10,True,tc);    ws1.cell(r,1).fill = F(bg)
    ws1.cell(r,2).value = "(sample)"; ws1.cell(r,2).font = fnt(10,False,tc);   ws1.cell(r,2).fill = F(bg); ws1.cell(r,2).alignment = aln("center")
    ws1.cell(r,4).value = note;       ws1.cell(r,4).font = fnt(9,False,MUTED); ws1.cell(r,4).fill = F(bg); ws1.cell(r,4).alignment = aln()

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — RESOURCE PLAN
# Col  A     = Role name
# Cols B–K   = Wk 1 – Wk 10   (10 project weeks)
# Cols L–M   = HC-Wk1, HC-Wk2  (hypercare)
# Col  N     = Total Days (SUM B:M)
# Col  O     = FTE Equiv (N/50)
# Col  P     = Phase focus note
#
# Rows 4–11  = 8 roles
# Row  12    = Total team-days per week (SUM rows 4:11)
# Row  13    = Max capacity (40 = 8 roles × 5d)
# Row  14    = Utilisation %
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resource Plan")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B4"
ws2.column_dimensions["A"].width = 26
for ci in range(2, 14): ws2.column_dimensions[get_column_letter(ci)].width = 7
ws2.column_dimensions["N"].width = 12
ws2.column_dimensions["O"].width = 11
ws2.column_dimensions["P"].width = 30

# Title
ws2.row_dimensions[1].height = 40
ws2.merge_cells("A1:P1")
ws2["A1"].value = "ANM Community App — Resource Loading Plan  (Team-Days per Week)"
ws2["A1"].font  = fnt(16, True, WHITE)
ws2["A1"].fill  = F(NAVY)
ws2["A1"].alignment = aln("left", indent=1)

# Row 2 — Phase header
ws2.row_dimensions[2].height = 20
hdr_cell(ws2, 2, 1, "PHASE", NAVY, WHITE, 9, True, "center", 0)
phase_spans = [
    ("Phase 1: Discovery & Design",  2,  3,  "1D4ED8"),
    ("Phase 2: Core Development",    4,  7,  "065F46"),
    ("Phase 3: QA & Testing",        8,  9,  "92400E"),
    ("Phase 4: Deployment",         10, 10,  "6B21A8"),
    ("Phase 5: Hypercare",          11, 13,  "991B1B"),
]
for ph, cs, ce, pc in phase_spans:
    sl, el = get_column_letter(cs), get_column_letter(ce)
    ws2.merge_cells(f"{sl}2:{el}2")
    ws2[f"{sl}2"].value = ph; ws2[f"{sl}2"].font = fnt(9,True,WHITE)
    ws2[f"{sl}2"].fill  = F(pc); ws2[f"{sl}2"].alignment = aln("center")
for c, lbl_ in [(14,"TOTAL"),(15,"FTE"),(16,"Phase Focus")]:
    ws2.cell(2,c).value = lbl_; ws2.cell(2,c).font = fnt(9,True,WHITE)
    ws2.cell(2,c).fill  = F(NAVY); ws2.cell(2,c).alignment = aln("center")

# Row 3 — Week labels
ws2.row_dimensions[3].height = 18
hdr_cell(ws2, 3, 1, "Role", NAVY, WHITE, 10, True, "left", 1)
weeks12 = ["Wk 1","Wk 2","Wk 3","Wk 4","Wk 5","Wk 6","Wk 7","Wk 8","Wk 9","Wk 10","HC-1","HC-2"]
for ci, wl in enumerate(weeks12, 2):
    ws2.cell(3,ci).value = wl; ws2.cell(3,ci).font = fnt(9,True,WHITE)
    ws2.cell(3,ci).fill  = F(NAVY); ws2.cell(3,ci).alignment = aln("center")
for c, lbl_ in [(14,"Days"),(15,"÷50d"),(16,"Notes")]:
    ws2.cell(3,c).value = lbl_; ws2.cell(3,c).font = fnt(9,True,WHITE)
    ws2.cell(3,c).fill  = F(NAVY); ws2.cell(3,c).alignment = aln("center")

# Rows 4-11 — Resource allocations
# columns: role_name, wk1..wk10, hc1, hc2, phase_note
alloc = [
    ("Project Manager",          5,5,3,3,3,3,3,3,5,4,2,2, "Discovery → Go-Live → Hypercare oversight"),
    ("Solution Architect",       5,5,1,1,1,1,1,1,1,0,0,0, "Discovery, architecture gate & QA gate only"),
    ("UI/UX Designer",           5,5,2,2,1,1,0,0,0,0,0,0, "Design-heavy in Wks 1–4; wraps up Wk 6"),
    ("Mobile Developer (Sr.)",   0,1,5,5,5,5,3,3,4,2,2,2, "Full dev sprints 2–5; bug-fix in QA & HC"),
    ("Backend Developer (Sr.)",  2,2,5,5,5,5,3,3,3,1,2,2, "API design Wks 1–2; full from Wk 3 onward"),
    ("DevOps Engineer",          3,3,2,2,2,2,2,2,5,4,3,3, "Env setup Wks 1–2; full prod deploy Wk 9"),
    ("QA Engineer",              0,0,1,1,1,1,5,5,3,1,1,1, "Light in dev; full engagement Wks 7–8"),
    ("Security Consultant",      0,0,0,0,0,0,4,4,0,0,0,0, "OWASP pentest Wks 7–8 only; written report"),
]
for ri, row_data in enumerate(alloc):
    r = 4 + ri
    rname = row_data[0]; days = row_data[1:13]; note = row_data[13]
    ws2.row_dimensions[r].height = 20
    bg = WHITE if ri%2 == 0 else BG_ALT
    # Role name
    ws2.cell(r,1).value = rname; ws2.cell(r,1).font = fnt(10,True,NAVY)
    ws2.cell(r,1).fill  = F(bg); ws2.cell(r,1).alignment = aln()
    ws2.cell(r,1).border = btm_border()
    # Day allocations B-M (blue = inputs)
    for ci, d in enumerate(days, 2):
        cl = ws2.cell(r,ci)
        cl.value = d if d > 0 else None
        cl.font  = fnt(10, False, BLUE)
        cl.fill  = F(bg)
        cl.alignment = aln("center")
        cl.number_format = "#,##0;[Red]-#,##0;-"
        cl.border = Border(bottom=Side(style="thin",color="DDDDDD"),
                           left =Side(style="thin",color="E0E0E0"),
                           right=Side(style="thin",color="E0E0E0"))
    # N: Total days
    calc(ws2, r, 14, f"=SUM(B{r}:M{r})", "#,##0", BLACK, BG_HDR, "center", 10, True)
    ws2.cell(r,14).border = btm_border()
    # O: FTE equiv = total / 50  (10 weeks × 5d = 50 working days = 1 FTE)
    calc(ws2, r, 15, f"=N{r}/50", "0.0x", BLACK, BG_HDR, "center")
    ws2.cell(r,15).border = btm_border()
    # P: Notes
    ws2.cell(r,16).value = note; ws2.cell(r,16).font = fnt(9,False,MUTED)
    ws2.cell(r,16).fill  = F(bg); ws2.cell(r,16).alignment = aln()
    ws2.cell(r,16).border = btm_border()

# Row 12 — Total team-days per week (SUM rows 4:11)
ws2.row_dimensions[12].height = 22
ws2.cell(12,1).value = "TOTAL TEAM-DAYS / WEEK"
ws2.cell(12,1).font  = fnt(10,True,WHITE); ws2.cell(12,1).fill = F(NAVY); ws2.cell(12,1).alignment = aln(indent=1)
for ci in range(2, 15):
    c = get_column_letter(ci)
    calc(ws2, 12, ci, f"=SUM({c}4:{c}11)", "#,##0", WHITE, NAVY, "center", 10, True)

# Row 13 — Max capacity (8 × 5 = 40 per week)
ws2.row_dimensions[13].height = 17
ws2.cell(13,1).value = "Max capacity (8 roles × 5d)"; ws2.cell(13,1).font = fnt(9,False,MUTED); ws2.cell(13,1).alignment = aln()
for ci in range(2,14):
    inp(ws2, 13, ci, 40, "#,##0", h="center"); ws2.cell(13,ci).font = fnt(9,False,MUTED)
calc(ws2, 13, 14, "=SUM(B13:M13)", "#,##0", MUTED, h="center")

# Row 14 — Utilisation %
ws2.row_dimensions[14].height = 17
ws2.cell(14,1).value = "Utilisation %"; ws2.cell(14,1).font = fnt(9,False,MUTED); ws2.cell(14,1).alignment = aln()
for ci in range(2,14):
    c = get_column_letter(ci)
    calc(ws2, 14, ci, f'=IFERROR({c}12/{c}13,"-")', "0%", MUTED, h="center", sz=9)

# Phase summary table (rows 16-24)
ws2.row_dimensions[16].height = 8
ws2.row_dimensions[17].height = 20
ws2.merge_cells("A17:E17")
hdr_cell(ws2, 17, 1, "  PHASE SUMMARY")
ws2.row_dimensions[18].height = 17
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"% of Total"),(5,"Avg daily headcount")]:
    ws2.cell(18,c).value = h; ws2.cell(18,c).font = fnt(9,True,NAVY)
    ws2.cell(18,c).fill  = F(BG_HDR); ws2.cell(18,c).alignment = aln("center")
ph_summ = [
    ("Phase 1: Discovery & Design",   "Wks 1–2",   "=SUM(B12:C12)",  2),
    ("Phase 2: Core Development",     "Wks 3–6",   "=SUM(D12:G12)",  4),
    ("Phase 3: QA & Testing",         "Wks 7–8",   "=SUM(H12:I12)",  2),
    ("Phase 4: Deployment",           "Wk 9",      "=J12",           1),
    ("Phase 5: Hypercare (Wk10+HC)",  "Wks 10–12", "=SUM(K12:M12)",  3),
    ("TOTAL",                         "12 wks",    "=SUM(B12:M12)",   0),
]
for i, (ph, wk, days_f, n_wks) in enumerate(ph_summ):
    r = 19 + i; is_tot = (i==5)
    bg2 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT); tc2 = WHITE if is_tot else NAVY
    ws2.row_dimensions[r].height = 17
    ws2.cell(r,1).value = ph; ws2.cell(r,1).font = fnt(9,is_tot,tc2); ws2.cell(r,1).fill = F(bg2); ws2.cell(r,1).alignment = aln()
    ws2.cell(r,2).value = wk; ws2.cell(r,2).font = fnt(9,is_tot,tc2); ws2.cell(r,2).fill = F(bg2); ws2.cell(r,2).alignment = aln("center")
    calc(ws2, r, 3, days_f, "#,##0", tc2, bg2, "center", 9, is_tot)
    calc(ws2, r, 4, f'=IFERROR(C{r}/C{19+5},"-")', PCT, tc2, bg2, "center", 9)
    # avg daily headcount = team-days / (weeks × 5)
    wk_count = n_wks if n_wks > 0 else 12
    calc(ws2, r, 5, f'=IFERROR(ROUND(C{r}/({wk_count}*5),1),"-")', "0.0", tc2, bg2, "center", 9)
    for c in range(1,6): ws2.cell(r,c).border = btm_border()

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — COST MODEL
# Rows 3     = headers
# Rows 4-11  = 8 roles
# Row  12    = Subtotal (total cost, no margin)
# Row  14    = Total Cost USD (= H12)
# Row  15    = Margin % (= Assumptions!B25)
# Row  16    = Margin $ (= H14 × H15)
# Row  17    = Client Price (= H14 + H16)
# Rows 20+   = Phase cost breakdown
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Model")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A4"
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 15
ws3.column_dimensions["D"].width = 13
ws3.column_dimensions["E"].width = 13
ws3.column_dimensions["F"].width = 18
ws3.column_dimensions["G"].width = 16
ws3.column_dimensions["H"].width = 18
ws3.column_dimensions["I"].width = 11
ws3.column_dimensions["J"].width = 28

# Title
ws3.row_dimensions[1].height = 40
ws3.merge_cells("A1:J1")
ws3["A1"].value = "ANM Community App — Cost Build-up & Pricing Model"
ws3["A1"].font  = fnt(16,True,WHITE); ws3["A1"].fill = F(NAVY); ws3["A1"].alignment = aln("left",indent=1)

ws3.row_dimensions[2].height = 14
ws3.merge_cells("A2:J2")
ws3["A2"].value = "Green = cross-sheet link  ·  Blue = editable input  ·  Black = formula  ·  Margin applied in rows 14–17  ·  Amounts in USD"
ws3["A2"].font  = fnt(8,False,MUTED); ws3["A2"].fill = F(BG_HDR); ws3["A2"].alignment = aln("left",indent=1)

# Column headers row 3
ws3.row_dimensions[3].height = 30
col3_hdrs = ["Role","Daily Rate (USD)","Project Days\n(Wks 1–10)","HC Days\n(2 wks)",
             "Total\nDays","Project Cost\n(USD)","HC Cost\n(USD)","Total Cost\n(USD)","% of\nTotal","Source / Notes"]
for ci, h in enumerate(col3_hdrs, 1):
    ws3.cell(3,ci).value = h; ws3.cell(3,ci).font = fnt(9,True,WHITE); ws3.cell(3,ci).fill = F(NAVY)
    ws3.cell(3,ci).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Rows 4-11 — Role cost rows
# Resource Plan resource rows: 4=PM, 5=Arch, 6=Des, 7=Mob, 8=Back, 9=Dev, 10=QA, 11=Sec
# Assumptions rate rows:      B13=PM,B14=Arch,B15=Des,B16=Mob,B17=Back,B18=Dev,B19=QA,B20=Sec
# Resource Plan cols:         B-K = Wk1-Wk10 (proj), L-M = HC weeks
role_notes = [
    "Source: Assumptions!B13  |  Days: Resource Plan rows 4, cols B:K and L:M",
    "Source: Assumptions!B14  |  Primarily Wks 1–2; gates in later phases",
    "Source: Assumptions!B15  |  Design-heavy Wks 1–4; wraps by Wk 6",
    "Source: Assumptions!B16  |  Full sprints Wks 3–8; part-time in HC",
    "Source: Assumptions!B17  |  Full sprints Wks 3–8; part-time in HC",
    "Source: Assumptions!B18  |  Env setup Wks 1–2; full deployment Wk 9",
    "Source: Assumptions!B19  |  Full engagement Wks 7–8 QA phase",
    "Source: Assumptions!B20  |  OWASP pentest Wks 7–8 only; fixed scope",
]
role_data_rows = []
for ri in range(8):
    r = 4 + ri
    rp_row  = 4 + ri      # Resource Plan row for this role
    ar_row  = 13 + ri     # Assumptions rate row for this role
    role_data_rows.append(r)
    ws3.row_dimensions[r].height = 20
    bg = WHITE if ri%2==0 else BG_ALT
    # A: Role name (pulled from alloc list defined above)
    ws3.cell(r,1).value = alloc[ri][0]; ws3.cell(r,1).font = fnt(10,True,NAVY)
    ws3.cell(r,1).fill  = F(bg); ws3.cell(r,1).alignment = aln()
    # B: Daily rate (green link to Assumptions)
    xlink(ws3, r, 2, f"=Assumptions!B{ar_row}", "$#,##0", "center", bg)
    # C: Project days Wks 1-10 (green link — Resource Plan cols B:K)
    xlink(ws3, r, 3, f"=SUM('Resource Plan'!B{rp_row}:K{rp_row})", "#,##0", "center", bg)
    # D: Hypercare days (green link — Resource Plan cols L:M)
    xlink(ws3, r, 4, f"=SUM('Resource Plan'!L{rp_row}:M{rp_row})", "#,##0", "center", bg)
    # E: Total days = C + D
    calc(ws3, r, 5, f"=C{r}+D{r}", "#,##0", BLACK, bg, "center")
    # F: Project cost = B × C
    calc(ws3, r, 6, f"=B{r}*C{r}", MONEY, BLACK, bg)
    # G: HC cost = B × D
    calc(ws3, r, 7, f"=B{r}*D{r}", MONEY, BLACK, bg)
    # H: Total cost = F + G
    calc(ws3, r, 8, f"=F{r}+G{r}", MONEY, BLACK, bg)
    # I: % of total (reference H12 subtotal — formula added after subtotal row)
    ws3.cell(r,9).font = fnt(10,False,BLACK); ws3.cell(r,9).fill = F(bg)
    ws3.cell(r,9).alignment = aln("center"); ws3.cell(r,9).number_format = PCT
    # J: Note
    ws3.cell(r,10).value = role_notes[ri]; ws3.cell(r,10).font = fnt(8,False,MUTED); ws3.cell(r,10).alignment = aln()
    for c in range(1,10): ws3.cell(r,c).border = btm_border()

# Row 12 — Subtotal (total cost before margin)
SUB = 12
ws3.row_dimensions[SUB].height = 22
hdr_cell(ws3, SUB, 1, "  TOTAL COST  (before margin)", NAVY, WHITE, 10, True, "left", 1)
ws3.cell(SUB,2).fill = F(NAVY)
for ci, col in enumerate(["C","D","E","F","G","H"], 3):
    fmt_ = MONEY if ci >= 6 else "#,##0"
    h_   = "right" if ci >= 6 else "center"
    calc(ws3, SUB, ci, f"=SUM({col}4:{col}11)", fmt_, WHITE, NAVY, h_, 10, True)
ws3.cell(SUB,9).value = "100.0%"; ws3.cell(SUB,9).font = fnt(10,True,WHITE)
ws3.cell(SUB,9).fill  = F(NAVY); ws3.cell(SUB,9).alignment = aln("center")

# Fill % of total for each role row now that SUB row is known
for r in role_data_rows:
    ws3.cell(r,9).value = f"=IFERROR(H{r}/H{SUB},0)"

# ── Pricing Block (rows 14-17) ────────────────────────────
ws3.row_dimensions[13].height = 8   # gap

# Row 14: Total Cost = H12
# Row 15: Margin % = Assumptions!B25
# Row 16: Margin $ = H14 × H15
# Row 17: Client Price = H14 + H16
PSTART = 14
pricing_rows = [
    (14, "Total Cost (USD)",          f"=H{SUB}",                   F(BG_HDR),    fnt(11,True,NAVY),          MONEY,  "right"),
    (15, "Gross Margin % (50%)",       "=Assumptions!B25",           F("FEF9C3"),  fnt(11,True,BLUE),          "0.0%", "right"),
    (16, "Margin Amount (USD)",        f"=H{PSTART}*H{PSTART+1}",   F("F0FDF4"),  fnt(11,True,"065F46"),      MONEY,  "right"),
    (17, "TOTAL CLIENT PRICE (USD)",   f"=H{PSTART}+H{PSTART+2}",  F(GOLD),      fnt(13,True,NAVY),          MONEY,  "right"),
]
for (rnum, label, val, bg_fill_, txt_fnt, fmt_, h_) in pricing_rows:
    ws3.row_dimensions[rnum].height = 26
    ws3.merge_cells(f"A{rnum}:G{rnum}")
    ws3.cell(rnum,1).value = label; ws3.cell(rnum,1).font = txt_fnt
    ws3.cell(rnum,1).fill  = bg_fill_; ws3.cell(rnum,1).alignment = aln("left", indent=2)
    ws3.cell(rnum,8).value = val;   ws3.cell(rnum,8).font = txt_fnt
    ws3.cell(rnum,8).fill  = bg_fill_; ws3.cell(rnum,8).alignment = aln(h_)
    ws3.cell(rnum,8).number_format = fmt_
    ws3.cell(rnum,8).border = Border(bottom=Side(style="medium",color="AAAAAA"), top=Side(style="thin",color="CCCCCC"))

# ── Phase cost breakdown (rows 19+) ──────────────────────
ws3.row_dimensions[18].height = 8
PHSTART3 = 19
ws3.row_dimensions[PHSTART3].height = 22
ws3.merge_cells(f"A{PHSTART3}:J{PHSTART3}")
hdr_cell(ws3, PHSTART3, 1, "  COST & PRICE BY DELIVERY PHASE")

ws3.row_dimensions[PHSTART3+1].height = 28
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"Proportional Cost (USD)"),(5,"Selling Price\nincl. 50% Margin"),(6,"% of Total Price")]:
    ws3.cell(PHSTART3+1,c).value = h; ws3.cell(PHSTART3+1,c).font = fnt(9,True,NAVY)
    ws3.cell(PHSTART3+1,c).fill  = F(BG_HDR); ws3.cell(PHSTART3+1,c).alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

# Phase cost = (phase team-days / total team-days) × total cost
# Phase team-days: from Resource Plan row 12 (weekly totals)
ph_cost_data = [
    ("Phase 1: Discovery & Design",   "Wks 1–2",   "=SUM('Resource Plan'!B12:C12)"),
    ("Phase 2: Core Development",     "Wks 3–6",   "=SUM('Resource Plan'!D12:G12)"),
    ("Phase 3: QA & Testing",         "Wks 7–8",   "=SUM('Resource Plan'!H12:I12)"),
    ("Phase 4: Deployment",           "Wk 9",      "='Resource Plan'!J12"),
    ("Phase 5: Hypercare (Wk10+HC)",  "Wks 10–12", "=SUM('Resource Plan'!K12:M12)"),
    ("TOTAL",                         "12 wks",    f"=E{SUB}"),
]
for i, (ph, wk, days_f) in enumerate(ph_cost_data):
    r = PHSTART3 + 2 + i
    ws3.row_dimensions[r].height = 18
    is_tot = (i == 5)
    bg3 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT)
    tc3 = WHITE if is_tot else NAVY
    ws3.cell(r,1).value = ph; ws3.cell(r,1).font = fnt(9,is_tot,tc3); ws3.cell(r,1).fill = F(bg3); ws3.cell(r,1).alignment = aln()
    ws3.cell(r,2).value = wk; ws3.cell(r,2).font = fnt(9,is_tot,tc3); ws3.cell(r,2).fill = F(bg3); ws3.cell(r,2).alignment = aln("center")
    # Team-days: green link if not total row
    if not is_tot:
        xlink(ws3, r, 3, days_f, "#,##0", "center", bg3)
    else:
        calc(ws3, r, 3, days_f, "#,##0", WHITE, NAVY, "center", 9, True)
    # Cost = proportional: (phase_days / total_days) × total_cost
    calc(ws3, r, 4, f"=IFERROR(C{r}/E{SUB}*H{SUB},0)", MONEY, tc3, bg3, "right", 9, is_tot)
    # Selling price = phase_cost × (1 + margin)
    sell_tc = GOLD if is_tot else "065F46"; sell_bg = GOLD if is_tot else bg3
    calc(ws3, r, 5, f"=IFERROR(D{r}*(1+Assumptions!B25),0)", MONEY, sell_tc, sell_bg, "right", 9, is_tot)
    # % of total client price
    client_price_row = PSTART + 3   # row 17
    calc(ws3, r, 6, f'=IFERROR(E{r}/H{client_price_row},"-")', PCT, tc3, bg3, "center", 9)
    for c in range(1,7): ws3.cell(r,c).border = btm_border()

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — WEEKLY COST
# Row  3    = headers
# Rows 4-15 = 12 weeks (Wk1–Wk10, HC-Wk1, HC-Wk2)
# Row  16   = TOTAL
# Rows 18+  = Phase summary
#
# Cols: A=Week, B=Phase, C-J=8 role costs, K=Total/wk,
#       L=Cumul Cost, M=Sell Price/wk, N=Cumul Sell Price
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Weekly Cost")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "C4"
ws4.column_dimensions["A"].width = 9
ws4.column_dimensions["B"].width = 22
for ci in range(3, 11): ws4.column_dimensions[get_column_letter(ci)].width = 13
ws4.column_dimensions["K"].width = 16
ws4.column_dimensions["L"].width = 17
ws4.column_dimensions["M"].width = 16
ws4.column_dimensions["N"].width = 18

# Title
ws4.row_dimensions[1].height = 40
ws4.merge_cells("A1:N1")
ws4["A1"].value = "ANM Community App — Weekly Cost Breakdown & Cash Flow"
ws4["A1"].font  = fnt(16,True,WHITE); ws4["A1"].fill = F(NAVY); ws4["A1"].alignment = aln("left",indent=1)

ws4.row_dimensions[2].height = 14
ws4.merge_cells("A2:N2")
ws4["A2"].value = "Costs = days (Resource Plan) × rate (Assumptions)  ·  Selling price = cost × (1 + 50% margin)  ·  Amounts in USD"
ws4["A2"].font  = fnt(8,False,MUTED); ws4["A2"].fill = F(BG_HDR); ws4["A2"].alignment = aln("left",indent=1)

# Column headers row 3
ws4.row_dimensions[3].height = 30
role_short = ["PM","Architect","Designer","Mobile Dev","Backend Dev","DevOps","QA","Security"]
col4_hdrs  = ["Week","Phase"] + role_short + ["Total Cost\n/Wk (USD)","Cumul. Cost\n(USD)","Sell Price\n/Wk (USD)","Cumul. Sell\nPrice (USD)"]
for ci, h in enumerate(col4_hdrs, 1):
    ws4.cell(3,ci).value = h; ws4.cell(3,ci).font = fnt(9,True,WHITE); ws4.cell(3,ci).fill = F(NAVY)
    ws4.cell(3,ci).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wk_phase_info = [
    ("Wk 1",   "Phase 1: Discovery & Design", "EFF6FF"),
    ("Wk 2",   "Phase 1: Discovery & Design", "EFF6FF"),
    ("Wk 3",   "Phase 2: Core Development",   "F0FDF4"),
    ("Wk 4",   "Phase 2: Core Development",   "F0FDF4"),
    ("Wk 5",   "Phase 2: Core Development",   "F0FDF4"),
    ("Wk 6",   "Phase 2: Core Development",   "F0FDF4"),
    ("Wk 7",   "Phase 3: QA & Testing",       "FFFBEB"),
    ("Wk 8",   "Phase 3: QA & Testing",       "FFFBEB"),
    ("Wk 9",   "Phase 4: Deployment",         "F5F3FF"),
    ("Wk 10",  "Phase 5: Hypercare",          "FFF1F2"),
    ("HC-Wk1", "Phase 5: Hypercare",          "FFF1F2"),
    ("HC-Wk2", "Phase 5: Hypercare",          "FFF1F2"),
]
# Resource Plan: Wk1=col B, Wk2=col C, ..., Wk10=col K, HC1=col L, HC2=col M
rp_week_cols = ["B","C","D","E","F","G","H","I","J","K","L","M"]
# Resource Plan role rows: PM=4 … Sec=11
rp_rows4 = [4,5,6,7,8,9,10,11]
# Assumptions rates:
assum_rates = [f"Assumptions!B{13+i}" for i in range(8)]

for wi, (wlbl, wphase, wbg) in enumerate(wk_phase_info):
    r = 4 + wi
    ws4.row_dimensions[r].height = 20
    rp_col = rp_week_cols[wi]
    # A: Week label
    ws4.cell(r,1).value = wlbl;   ws4.cell(r,1).font = fnt(10,True,NAVY);  ws4.cell(r,1).fill = F(wbg); ws4.cell(r,1).alignment = aln("center")
    # B: Phase
    ws4.cell(r,2).value = wphase; ws4.cell(r,2).font = fnt(9,False,MUTED); ws4.cell(r,2).fill = F(wbg); ws4.cell(r,2).alignment = aln()
    # C-J: Role cost = Resource Plan days × rate
    for ci_off, (rp_row, rate_ref) in enumerate(zip(rp_rows4, assum_rates)):
        ci = 3 + ci_off
        calc(ws4, r, ci, f"='Resource Plan'!{rp_col}{rp_row}*{rate_ref}", MONEY, BLACK, wbg)
        ws4.cell(r,ci).border = btm_border()
    # K: Total cost this week
    calc(ws4, r, 11, f"=SUM(C{r}:J{r})", MONEY, BLACK, BG_HDR, "right", 10, True)
    ws4.cell(r,11).border = Border(bottom=Side(style="thin",color="DDDDDD"), left=Side(style="medium",color=NAVY))
    # L: Cumulative cost
    if wi == 0:
        calc(ws4, r, 12, f"=K{r}", MONEY, BLACK, BG_HDR, "right")
    else:
        calc(ws4, r, 12, f"=L{r-1}+K{r}", MONEY, BLACK, BG_HDR, "right")
    ws4.cell(r,12).border = btm_border()
    # M: Selling price this week = K × (1 + margin)
    calc(ws4, r, 13, f"=K{r}*(1+Assumptions!B25)", MONEY, "065F46", "F0FDF4", "right", 10, True)
    ws4.cell(r,13).border = Border(bottom=Side(style="thin",color="DDDDDD"), left=Side(style="medium",color=NAVY))
    # N: Cumulative selling price
    if wi == 0:
        calc(ws4, r, 14, f"=M{r}", MONEY, "065F46", "F0FDF4", "right")
    else:
        calc(ws4, r, 14, f"=N{r-1}+M{r}", MONEY, "065F46", "F0FDF4", "right")
    ws4.cell(r,14).border = btm_border()

# Row 16 — Grand total
GTROW = 16
ws4.row_dimensions[GTROW].height = 24
ws4.cell(GTROW,1).value = "TOTAL"; ws4.cell(GTROW,1).font = fnt(11,True,WHITE); ws4.cell(GTROW,1).fill = F(NAVY); ws4.cell(GTROW,1).alignment = aln("center")
ws4.cell(GTROW,2).fill = F(NAVY)
for ci in range(3, 15):
    c = get_column_letter(ci)
    if ci in [12, 14]:                          # Cumulative: show final value
        calc(ws4, GTROW, ci, f"={c}{GTROW-1}", MONEY, NAVY, GOLD, "right", 12, True)
    else:                                       # Summable: sum all 12 weeks
        calc(ws4, GTROW, ci, f"=SUM({c}4:{c}{GTROW-1})", MONEY, WHITE, NAVY, "right", 11, True)

# Phase summary (rows 18+)
ws4.row_dimensions[18].height = 8
PHROW4 = 19
ws4.row_dimensions[PHROW4].height = 22
ws4.merge_cells(f"A{PHROW4}:N{PHROW4}")
hdr_cell(ws4, PHROW4, 1, "  COST SUMMARY BY PHASE")
ws4.row_dimensions[PHROW4+1].height = 18
for c, h in [(1,"Phase"),(2,"Weeks"),(3,"Team-Days"),(4,"Cost (USD)"),(5,"Selling Price (USD)"),(6,"% of Total")]:
    ws4.cell(PHROW4+1,c).value = h; ws4.cell(PHROW4+1,c).font = fnt(9,True,NAVY)
    ws4.cell(PHROW4+1,c).fill  = F(BG_HDR); ws4.cell(PHROW4+1,c).alignment = aln("center")
# Phase ranges in Weekly Cost (K = total cost/wk, M = sell/wk)
ph4_ranges = [
    ("Phase 1: Discovery & Design",   "Wks 1–2",   "K4:K5",   "M4:M5",   "=SUM('Resource Plan'!B12:C12)"),
    ("Phase 2: Core Development",     "Wks 3–6",   "K6:K9",   "M6:M9",   "=SUM('Resource Plan'!D12:G12)"),
    ("Phase 3: QA & Testing",         "Wks 7–8",   "K10:K11", "M10:M11", "=SUM('Resource Plan'!H12:I12)"),
    ("Phase 4: Deployment",           "Wk 9",      "K12",     "M12",     "='Resource Plan'!J12"),
    ("Phase 5: Hypercare (Wk10+HC)",  "Wks 10–12", "K13:K15", "M13:M15", "=SUM('Resource Plan'!K12:M12)"),
    ("TOTAL",                         "12 wks",    "K4:K15",  "M4:M15",  "=SUM('Resource Plan'!B12:M12)"),
]
TOTAL_SELL_ROW = PHROW4 + 2 + 5   # the TOTAL phase row
for i, (ph, wk, kr, mr, days_f) in enumerate(ph4_ranges):
    r = PHROW4 + 2 + i
    ws4.row_dimensions[r].height = 18
    is_tot = (i == 5)
    bg4 = NAVY if is_tot else (WHITE if i%2==0 else BG_ALT); tc4 = WHITE if is_tot else NAVY
    ws4.cell(r,1).value = ph; ws4.cell(r,1).font = fnt(9,is_tot,tc4); ws4.cell(r,1).fill = F(bg4); ws4.cell(r,1).alignment = aln()
    ws4.cell(r,2).value = wk; ws4.cell(r,2).font = fnt(9,is_tot,tc4); ws4.cell(r,2).fill = F(bg4); ws4.cell(r,2).alignment = aln("center")
    xlink(ws4, r, 3, days_f, "#,##0", "center", bg4) if not is_tot else calc(ws4, r, 3, days_f, "#,##0", WHITE, NAVY, "center", 9, True)
    calc(ws4, r, 4, f"=SUM({kr})", MONEY, tc4, bg4, "right", 9, is_tot)
    calc(ws4, r, 5, f"=SUM({mr})", MONEY, GOLD if is_tot else "065F46", GOLD if is_tot else bg4, "right", 9, is_tot)
    calc(ws4, r, 6, f'=IFERROR(E{r}/E{TOTAL_SELL_ROW},"-")', PCT, tc4, bg4, "center", 9)
    for c in range(1,7): ws4.cell(r,c).border = btm_border()

# ── Save ──────────────────────────────────────────────────
out_path = "ANM_Resource_Cost_Model.xlsx"
wb.save(out_path)
print(f"\n✅  {out_path} generated successfully!")
print("   Open in Excel or LibreOffice Calc.")
print("\n   Key outputs (approximate):")
print("     Total Cost  : ~$193,200")
print("     50% Margin  : ~$96,600")
print("     Client Price: ~$289,800")
print("     Total Days  : ~210 team-days across 12 weeks")
